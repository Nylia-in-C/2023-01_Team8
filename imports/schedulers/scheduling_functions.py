import os, sys
import math
import datetime as dt
import pandas as pd
import holidays
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl import Workbook

import helpers

currentdir = os.path.dirname(os.path.realpath(__file__))
parentdir = os.path.dirname(currentdir)
grandparentdir = os.path.dirname(parentdir)
sys.path.append(grandparentdir)

from imports.classes.courses import *
from imports.classes.classrooms import *
from imports.fillClassrooms import *
from database.database import *
from itertools import cycle
from collections import defaultdict
from typing import *
# from openpyxl.utils import get_column_letter
# from openpyxl import Workbook


#=========================== INITIAL DATA RETREIVAL ============================
def get_courses(prog: str, term: int) -> List[Course]:
    ''' fetch all courses for a given program & term from the database '''
    db = helpers.check_path("database\database.db")  # database.db file path
    connection = create_connection(db)

    query = f"SELECT C.* FROM \
              Courses C JOIN Programs P ON C.CourseID = P.CourseID WHERE \
              P.ProgID = '{prog}' AND C.Term = {term};"

    try:
        cur = connection.cursor()
        cur.execute(query)
        rows = cur.fetchall()
    except:
        pass
        #print("unable to retrieve core courses from database")
        return

    close_connection(connection)

    # convert strings from the database into Course objects
    courses = []
    for row in rows:
        row = list(row)
        # convert 0/1 to booleans
        for i in range(5, 8):
            if   row[i] == 0: row[i] = False
            elif row[i] == 1: row[i] = True

        courses.append( Course(*row) )

    return courses
    
def get_rooms() -> List[Classroom]:
    ''' fetch all non-ghost classrooms from the database'''
    db = helpers.check_path("database\database.db")  # database.db file path
    connection = create_connection(db)

    query = f"SELECT * FROM Classrooms C;" #WHERE C.ClassID NOT LIKE 'ghost%';"

    try:
        cur = connection.cursor()
        cur.execute(query)
        rows = cur.fetchall()
    except:
        pass
        #print("unable to retrieve classrooms from database")

    close_connection(connection)

    rooms = []
    for row in rows:
        row = list(row)
        isLab = False
        if int(row[2]) == 1:
            isLab = True
        rooms.append(Classroom(row[0], int(row[1]), isLab))

    return rooms

def get_cohort_counts(term: int) -> Dict[str, int]:
    ''' 
    call `fillClassrooms` to get a breakdown of all cohorts, then return a 
    cohort count for each program/term
    '''
    cohort_dict = fillClassrooms(term)
    counts = {}
    for program, cohorts in cohort_dict.items():
        if all(c == 0 for c in cohorts): 
            counts[program] = 0
        else:
            counts[program] = len(cohorts)
    
    return counts
#===============================================================================


#============================ HELPER FUNCTIONS =================================
def make_empty_scheds(room_list: List[Classroom], is_program: bool = False) -> pd.DataFrame:
    '''
    make an empty pandas dataframe to represent the schedule for a single day. 
    Column headers are room numbers & row indexes are times (half hour increments)
    '''
    # schedule for program-specific courses goes from 8:00 - 8:30 pm 
    # (4:30 - 8:30 is for fullstack)
    if is_program:
        times = [dt.time(i, j).strftime("%H:%M")
                 for i in range(8, 21) for j in [0, 30]]
    else:
        times = [dt.time(i, j).strftime("%H:%M")
                 for i in range(8, 17) for j in [0, 30]]
        times.append(dt.time(17, 0).strftime("%H:%M"))
    
    lec_sched = pd.DataFrame(index=times)
    lab_sched = pd.DataFrame(index=times)
    onl_sched = pd.DataFrame(index=times)
    
    for room in room_list:
        if room.isLab:
            lab_sched[f"{room.ID} (LAB)"] = [""] * len(lab_sched.index)
            
        elif room.ID == "ONLINE":
            onl_sched[room.ID] = [""] * len(onl_sched.index)
            
        else:
            lec_sched[room.ID] = [""] * len(lec_sched.index)
            
    return {"lecture": lec_sched, "lab": lab_sched, "online": onl_sched}

def get_course_hours(courses: List[Course]) -> Dict[str, int]:
    '''
    Create a dictionary that will keep track of how many hours
    a given course has, and how many hours it has left. 
    '''
    hours = {}

    for course in courses:
        hours[course.ID] = {
            'required': course.termHours,
            'remaining': course.termHours,
            'scheduled': 0,
        }

    return hours

def update_course_hours(course_hours: Dict[str, int], prev_schedule: pd.DataFrame) -> Dict[str, str]:
    '''
    Use the previous day's schedule to update the remaining hours for each course
    '''
    # dont update the same course twice
    seen = set()

    # get number of hours for each course on the schedule
    for room in prev_schedule.columns:
        courses = prev_schedule[room].tolist()
        for course in courses:
            if (course == ""):
                continue

            # remove cohort identifier from non-online courses
            course_name = course[:-3] if room != 'ONLINE' else course

            if course_name in seen:
                continue

            seen.add(course_name)
            
            scheduled_hours = (courses.count(course)) * 0.5
                
            course_hours[course_name]['scheduled'] += scheduled_hours
            course_hours[course_name]['remaining'] -= scheduled_hours

    return course_hours

def update_schedule(course_hours: Dict[str, int], prev_sched: pd.DataFrame,
                    date_ints: Dict[str, int], end_days: Dict[str, int]) -> Dict[str, pd.DataFrame]: 
    '''
    Takes the previous day's schedules and the course hours dict. If any courses
    have met their term hour requirements, they are removed and replaced 
    with empty strings the updated schedules are returned in a dict
    '''
    # list of courses in the last schedule with, and without their cohort IDs
    prev_courses = list((prev_sched.applymap(lambda x: x[:-3])).values.flatten()) +\
                   list(prev_sched.values.flatten())

    # courses that are in the schedule and have met their required hours should
    # be replaced with empty strings 
    remove = [id for id in course_hours if course_hours[id]['remaining'] <= 0
              and id in prev_courses]
    
    # update end dates for any courses that are removed
    for course in remove:
        end_days[course] = date_ints['day']
    
    # regex patterns for matching courses that will be removed 
    remove_pattern = r'^(' + '|'.join(remove) + ')(-\d{2})?$'
    
    # remove any finished courses from the schedule
    updated_sched = prev_sched.apply(lambda x: x.str.replace(remove_pattern, "", regex=True))
    
    # split updated schedule into lecture, lab, and online schedules
    lab_col_idx = (updated_sched.columns.str.endswith("(LAB)")).argmax()
    onl_col_idx = (updated_sched.columns.str.endswith("ONLINE")).argmax()
    
    lec_sched = updated_sched.iloc[:, :lab_col_idx]
    lab_sched = updated_sched.iloc[:, lab_col_idx:onl_col_idx]
    onl_sched = updated_sched.iloc[:, onl_col_idx:]
    
    return {"lecture": lec_sched, "lab": lab_sched, "online": onl_sched}

def filter_courses(courses: List[Course], sched: pd.DataFrame, 
                   course_hours: Dict[str, int], day_count: int) -> List[Course]:
    '''
    Take a list of courses and filter out ones that have already been added to 
    the schedule or have already met their term hour requirements. Sort the list
    of courses in descending order of total term hours before returning it
    '''

    # PCOM 0130 & PCOM 0140 dont start until halfway through the term
    if day_count < 13:
        courses = list(filter(lambda c: c.ID not in ["PCOM 0130", "PCOM 0140"], courses))
    
    # AVDM 0260 runs on the last 2 days of the term, for 3 hour sessions
    if day_count < 25:
        courses = list(filter(lambda c: c.ID != "AVDM 0260", courses))

        
    # given course list should only contain a single type of course
    # (online courses dont include cohort IDs)
    if all(c.isOnline for c in courses):
        scheduled = [c for c in sched.values.flatten()]
    else:
        scheduled = [c[:-3] for c in sched.values.flatten()]
    
    valid = [c for c in courses if c.ID not in scheduled 
             and course_hours[c.ID]['remaining'] > 0]

    valid.sort(key=lambda x: x.termHours, reverse=True)
    
    return valid

def get_time_slot_count(course: Course, course_hours: Dict[str, int]) -> int:
    '''
    Checks a given courses duration, and how many term hours it has left and
    returns the number of half-hour blocks it should be scheduled for
    '''
    hours_left = course_hours[course.ID]['remaining']
    
    if (hours_left >= course.duration):
        return int(course.duration // 0.5)
    else:
        return math.ceil(hours_left / 0.5)

def get_available_times(sched: pd.DataFrame, blocks: int, 
                        cohorts: int, is_fs: bool = False) -> Dict[str, List[int]]:
    '''
    Takes the current schedule, the number of half-hour blocks a given course 
    will take up, and the number of cohorts needed. Returns a dictionary mapping
    room names to a list of available start times (indexes). If there arent enough
    available times for all cohorts, returns an empty dict
    '''
    available = defaultdict(list)
    required_gap = [""] * blocks

    # use a sliding window to find an available time & keep track of how many you find
    # full-stack starts at 4:30, everything else at 8:00
    left = 0 if not is_fs else 17
    right = left + blocks
    
    count = 0

    # keep track of when the last added course ends to prevent overlapping courses
    prev_course_end = -1

    # iterate rooms, move to the next time when we cycle back around to the first room
    for room in cycle(sched.columns.tolist()):

        # at the last room & the last time slot
        if room == list(sched.columns)[-1] and \
            ((is_fs and right > 25) or (not is_fs and right > 18)):
            break

        # found an open time gap
        if list(sched[room][left:right]) == required_gap:
            # check if adding current course will overlap with the previous one
            if (left >= prev_course_end):
                available[room].append(left)
                prev_course_end = right
                count += 1

        # reached the last room, move to the next starting time
        if room == list(sched.columns)[-1]:
            left += 1
            right += 1

    # return empty dict if there aren't enough available times for all cohorts
    if count < cohorts:
        return {}

    return available
    
def get_valid_cohorts(invalid_courses: List[str], start: int, end: int,
                      cohorts: List[str], curr_room: str, sched: pd.DataFrame) -> Set[str]:
    '''
    Given a potential start time and duration for a course, returns a set of 
    unique cohort IDs that can be used (prevents scheduling conflicts). 
    '''
    # list of all times the course will occupy
    occupied_times = [t for t in range(start, end)]
    # get dict of rooms mapped to the row indexes of any incompatible courses
    # scheduled at any of the time slots the new course might occupy
    matching_courses = {}
    for room in sched.columns:

        if room == curr_room:
            continue

        matches = sched[room].apply(
            lambda x: any([c in str(x[:-3]) for c in invalid_courses])
        )
        times = list(matches[matches].index.values)
        overlap = [t for t in times if times.index(t) in occupied_times]

        matching_courses[room] = overlap

    invalid_cohorts = set()
    # for each occupied time for the new course, add the unavailable cohorts to a set
    for room, times in matching_courses.items():
        for time in times:
            invalid_cohorts.add(sched[room][time][-2:])
            
    return set(cohorts).difference(invalid_cohorts)

def add_lecture_to_db(lec: Lecture) -> None:
    '''
    Add a lecture object to the database so the UI can read & display them
    '''
    db = helpers.check_path("database\database.db")  # database.db file path
    connection = create_connection(db)
    addLectureItem(connection, lec)
    close_connection(connection)
    return

def export_to_excel(sched_dict: Dict[str, pd.DataFrame]) -> None:


    if not os.path.exists("Exported Schedule.xlsx"):
        wb = Workbook()
        wb.save("Exported Schedule.xlsx")

    with pd.ExcelWriter("Exported Schedule.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        wb = load_workbook("Exported Schedule.xlsx")

        for day, val in sched_dict.items():

            try:
                next_empty_row = wb[f"DAY {day}"].max_row
                next_empty_row = 0 if next_empty_row == 0 else next_empty_row + 2
            except:
                next_empty_row = 0

            val.to_excel(writer, sheet_name=f"DAY {day}", index=False, startrow=next_empty_row)
            
            #TODO: make each tab a week schedule, rather than a single day
            
            # set columns widths so schedule is readable
            curr_sheet = writer.sheets[f"DAY {day}"]

            for column in curr_sheet.columns:
                col_length = max(len(str(cell.value)) for cell in column)
                col_char   = get_column_letter(column[0].column)

                # set the column width based on the maximum cell width found
                curr_sheet.column_dimensions[col_char].width = col_length + 5
    return


# testing only
def is_invalid_core_sched(lectures, sched):

    pcomA_strs = [c.ID for c in lectures['pcomA']]
    pcomB_strs = [c.ID for c in lectures['pcomB']]
    bcomA_strs = [c.ID for c in lectures['bcomA']]
    bcomB_strs = [c.ID for c in lectures['bcomB']]

    tr_sched = sched.T
    for time in tr_sched.columns.tolist():
        pcomA = [c[-1] for c in tr_sched[time].tolist() if c[:-3] in pcomA_strs]
        pcomB = [c[-1] for c in tr_sched[time].tolist() if c[:-3] in pcomB_strs]
        bcomA = [c[-1] for c in tr_sched[time].tolist() if c[:-3] in bcomA_strs]
        bcomB = [c[-1] for c in tr_sched[time].tolist() if c[:-3] in bcomB_strs]

        if len(pcomA) != len(set(pcomA)) or len(pcomB) != len(set(pcomB)) or \
           len(bcomA) != len(set(bcomA)) or len(bcomB) != len(set(bcomB)):
            return True
    return False

def is_invalid_prgm_sched(lectures, sched):
    pmA_strs = [c.ID for c in lectures['pmA']]
    pmB_strs = [c.ID for c in lectures['pmB']]
    baA_strs = [c.ID for c in lectures['baA']]
    baB_strs = [c.ID for c in lectures['baB']]
    glmA_strs = [c.ID for c in lectures['glmA']]
    glmB_strs = [c.ID for c in lectures['glmB']]
    dxdA_strs = [c.ID for c in lectures['dxdA']]
    dxdB_strs = [c.ID for c in lectures['dxdB']]
    bkA_strs = [c.ID for c in lectures['bkA']]
    bkB_strs = [c.ID for c in lectures['bkB']]
    fsA_strs = [c.ID for c in lectures['fsA']]
    fsB_strs = [c.ID for c in lectures['fsB']]
    
    transposed = sched.T
    for time in transposed.columns:
        pmA = [c[-1] for c in transposed[time].tolist() if c[:-3] in pmA_strs]
        pmB = [c[-1] for c in transposed[time].tolist() if c[:-3] in pmB_strs]
        baA = [c[-1] for c in transposed[time].tolist() if c[:-3] in baA_strs]
        baB = [c[-1] for c in transposed[time].tolist() if c[:-3] in baB_strs]
        glmA = [c[-1] for c in transposed[time].tolist() if c[:-3] in glmA_strs]
        glmB = [c[-1] for c in transposed[time].tolist() if c[:-3] in glmB_strs]
        dxdA = [c[-1] for c in transposed[time].tolist() if c[:-3] in dxdA_strs]
        dxdB = [c[-1] for c in transposed[time].tolist() if c[:-3] in dxdB_strs]
        bkA = [c[-1] for c in transposed[time].tolist() if c[:-3] in bkA_strs]
        bkB = [c[-1] for c in transposed[time].tolist() if c[:-3] in bkB_strs]
        fsA = [c[-1] for c in transposed[time].tolist() if c[:-3] in fsA_strs]
        fsB = [c[-1] for c in transposed[time].tolist() if c[:-3] in fsB_strs]
        
        parsed_IDs = pmA + pmB + baA + baB + glmA + glmB + \
                 dxdA + dxdB + bkA + bkB + fsA + fsB
                 
        if len(parsed_IDs) != len(set(parsed_IDs)):
            return True
    return False
#===============================================================================


#============================= LECTURE SCHEDULING ==============================
def make_core_lecture_sched(lectures: Dict[str, List[Course]], 
                            cohorts: Dict[str, List[str]], c_hours: Dict[str, int], 
                            sched: pd.DataFrame, date_ints: Dict[str, int]) -> pd.DataFrame:
    '''
    Takes a dict of pcom & bcom lectures for terms A & B, a dict of course_hours, 
    and a room list to create a schedule for a single day and return it as a 
    DataFrame. Courses are scheduled by pcom/bcom and term, moving from rooms
    left to right, starting at 8 am.
    '''
    week = date_ints['week']
    day  = date_ints['day']

    # lists of course IDs used to ensure no scheduling conflicts occur
    pcomA_strs = [c.ID for c in lectures['pcomA']]
    pcomB_strs = [c.ID for c in lectures['pcomB']]
    bcomA_strs = [c.ID for c in lectures['bcomA']]
    bcomB_strs = [c.ID for c in lectures['bcomB']]

    # parse cohorts for each program/term
    pcomA_cohorts = cohorts['pcomA']
    pcomB_cohorts = cohorts['pcomB']
    bcomA_cohorts = cohorts['bcomA']
    bcomB_cohorts = cohorts['bcomB']

    # filter out courses that dont need to be scheduled
    pcomA = filter_courses(lectures['pcomA'], sched, c_hours, day)
    pcomB = filter_courses(lectures['pcomB'], sched, c_hours, day)
    bcomA = filter_courses(lectures['bcomA'], sched, c_hours, day)
    bcomB = filter_courses(lectures['bcomB'], sched, c_hours, day)
    
    
    if len(bcomB) > 0:
            sched = add_lec(bcomB[0], bcomB_cohorts, c_hours,
                            bcomB_strs, sched, week, day, "BCOM")
            
    if len(pcomB) > 0:
            sched = add_lec(pcomB[0], pcomB_cohorts, c_hours,
                            pcomB_strs, sched, week, day, "PCOM")
            
    if len(pcomA) > 0:
            sched = add_lec(pcomA[0], pcomA_cohorts, c_hours,
                            pcomA_strs, sched, week, day, "PCOM")
            
    if len(bcomA) > 0:
            sched = add_lec(bcomA[0], bcomA_cohorts, c_hours, 
                            bcomA_strs, sched, week, day, "BCOM")

    return sched

def make_prgm_lecture_sched(lectures: Dict[str, List[Course]], 
                            cohorts: Dict[str, List[str]], 
                            c_hours: Dict[str, int], sched: pd.DataFrame,
                            date_ints: Dict[str, int]) -> pd.DataFrame:
    '''
    Takes a dict of program lectures for terms A & B, a dict of course_hours, 
    and a room list to create a schedule for a single day and return it as a 
    DataFrame. Courses are scheduled by program and term, moving from rooms
    left to right, starting at 8 am.
    '''
    week = date_ints['week']
    day  = date_ints['day']

    # lists of course IDs used to ensure no scheduling conflicts occur
    pmA_strs  = [c.ID for c in lectures['pmA']]
    pmB_strs  = [c.ID for c in lectures['pmB']]
    baA_strs  = [c.ID for c in lectures['baA']]
    baB_strs  = [c.ID for c in lectures['baB']]
    glmA_strs = [c.ID for c in lectures['glmA']]
    glmB_strs = [c.ID for c in lectures['glmB']]
    dxdA_strs = [c.ID for c in lectures['dxdA']]
    dxdB_strs = [c.ID for c in lectures['dxdB']]
    bkA_strs  = [c.ID for c in lectures['bkA']]
    bkB_strs  = [c.ID for c in lectures['bkB']]
    fsA_strs  = [c.ID for c in lectures['fsA']]
    fsB_strs  = [c.ID for c in lectures['fsB']]

    # parse cohorts for each program/term
    pmA_cohorts  = cohorts['pmA']
    pmB_cohorts  = cohorts['pmB']
    baA_cohorts  = cohorts['baA']
    baB_cohorts  = cohorts['baB']
    glmA_cohorts = cohorts['glmA']
    glmB_cohorts = cohorts['glmB']
    dxdA_cohorts = cohorts['dxdA']
    dxdB_cohorts = cohorts['dxdB']
    bkA_cohorts  = cohorts['bkA']
    bkB_cohorts  = cohorts['bkB']
    fsA_cohorts  = cohorts['fsA']
    fsB_cohorts  = cohorts['fsB']

    # filter out courses that dont need scheduling
    pmA  = filter_courses(lectures['pmA'],  sched, c_hours, day)
    pmB  = filter_courses(lectures['pmB'],  sched, c_hours, day)
    baA  = filter_courses(lectures['baA'],  sched, c_hours, day)
    baB  = filter_courses(lectures['baB'],  sched, c_hours, day)
    glmA = filter_courses(lectures['glmA'], sched, c_hours, day)
    glmB = filter_courses(lectures['glmB'], sched, c_hours, day)
    dxdA = filter_courses(lectures['dxdA'], sched, c_hours, day)
    dxdB = filter_courses(lectures['dxdB'], sched, c_hours, day)
    bkA  = filter_courses(lectures['bkA'],  sched, c_hours, day)
    bkB  = filter_courses(lectures['bkB'],  sched, c_hours, day)
    fsA  = filter_courses(lectures['fsA'],  sched, c_hours, day)
    fsB  = filter_courses(lectures['fsB'],  sched, c_hours, day)
    
    if len(pmB) > 0:
            sched = add_lec(pmB[0], pmB_cohorts, c_hours,
                            pmB_strs, sched, week, day, "PM",)
    if len(glmB) > 0:
            sched = add_lec(glmB[0], glmB_cohorts, c_hours,
                            glmB_strs, sched, week, day, "GLM",)
    if len(bkB) > 0:
            sched = add_lec(bkB[0], bkB_cohorts, c_hours,
                            bkB_strs, sched, week, day, "BK",)
    if len(baB) > 0:
            sched = add_lec(baB[0], baB_cohorts, c_hours,
                            baB_strs, sched, week, day, "BA",)
    if len(pmA) > 0:
            sched = add_lec(pmA[0], pmA_cohorts, c_hours, 
                            pmA_strs, sched, week, day, "PM",)
    if len(baA) > 0:
            sched = add_lec(baA[0], baA_cohorts, c_hours,
                            baA_strs, sched, week, day, "BA",)
    if len(glmA) > 0:
            sched = add_lec(glmA[0], glmA_cohorts, c_hours, 
                            glmA_strs, sched, week, day, "GLM",)
    if len(dxdA) > 0:
            sched = add_lec(dxdA[0], dxdA_cohorts, c_hours, 
                            dxdA_strs, sched, week, day, "DXD",)
    if len(dxdB) > 0:
            sched = add_lec(dxdB[0], dxdB_cohorts, c_hours, 
                            dxdB_strs, sched, week, day, "DXD",)
    if len(bkA) > 0:
            sched = add_lec(bkA[0], bkA_cohorts, c_hours,
                            bkA_strs, sched, week, day, "BK",)
    if len(fsA) > 0:
            sched = add_lec(fsA[0], fsA_cohorts, c_hours, 
                            fsA_strs, sched, week, day, "FS", True)
    if len(fsB) > 0:
            sched = add_lec(fsB[0], fsB_cohorts, c_hours, 
                            fsB_strs, sched, week, day, "FS", True)
        
    return sched

def add_lec(course: Course, cohorts: List[str], course_hours: Dict[str, int], 
            invalid_courses: List[str], sched: pd.DataFrame,  week: int, 
            day: int, prgm_str: str, is_fs: bool = False) -> pd.DataFrame:
    '''
    Checks when/if a given course can be scheduled and if so, what order the 
    cohorts should be scheduled. Returns an updated schedule if the course can
    be added, otherwise this returns an unaltered schedule
    '''

    blocks     = get_time_slot_count(course, course_hours)
    open_times = get_available_times(sched, blocks, len(cohorts), is_fs)
        
    if (open_times == {}):
        return sched

    cohort_times = {}
    for room, times in open_times.items():
        # for each start time, find out which cohorts can be scheduled 
        for start in times:
            end = start + blocks
            valid = get_valid_cohorts(
                invalid_courses, start, end, cohorts, room, sched, 
            )
            # only add cohorts we havent seen yet
            for cohort_ID in valid.difference(set(cohort_times.keys())):
                # only add rooms & start times that have not been assigned yet
                if (room, start) not in cohort_times.values():
                    cohort_times[cohort_ID] = (room, start)
                
    # not all cohorts can fit in the schedule
    if len(cohort_times.keys()) < len(cohorts):
        return sched

    for ID, time_slot in cohort_times.items():

        room, start = time_slot
        room_index  = sched.columns.get_loc(room)
        start_label = list(sched.index)[start]
        full_cohort = f"{prgm_str}0{course.term}{ID}"

        # store the scheduled course's info in the database
        add_lecture_to_db(
            Lecture(
                course.ID, course.title, course.termHours, course.term,
                course.duration, course.isCore, course.isOnline, course.hasLab,
                course.preReqs, full_cohort, room, week, day, start_label
            )
        )

        course_strs = [course.ID + '-' + ID] * blocks
        sched.iloc[start:(start+blocks), room_index] = course_strs

    return sched
#===============================================================================
    
    
#=============================== LAB SCHEDULING ================================
def make_core_lab_sched(lectures: Dict[str, List[Course]], 
                        labs: Dict[str, List[Course]],
                        cohorts: Dict[str, List[str]], c_hours: Dict[str, int], 
                        lec_sched: pd.DataFrame, lab_sched: pd.DataFrame,
                        date_ints: Dict[str, int]) -> pd.DataFrame:
    '''
    Takes the existing lecture schedule and information on what labs to schedule,
    and tries to schedule each lab without any conflicts with already scheduled
    courses
    '''
    day  = date_ints['day']
    week = date_ints['week']
    # lists of lecture & lab IDs used to ensure no scheduling conflicts occur
    pcomA_strs = [c.ID for c in lectures['pcomA']]+[c.ID for c in labs['pcomA']] 
    pcomB_strs = [c.ID for c in lectures['pcomB']]+[c.ID for c in labs['pcomB']]
    bcomA_strs = [c.ID for c in lectures['bcomA']]+[c.ID for c in labs['bcomA']]
    bcomB_strs = [c.ID for c in lectures['bcomB']]+[c.ID for c in labs['bcomB']]
    
    # parse cohorts for each program/term
    pcomA_cohorts = cohorts['pcomA']
    pcomB_cohorts = cohorts['pcomB']
    bcomA_cohorts = cohorts['bcomA']
    bcomB_cohorts = cohorts['bcomB']
    
    # filter out labs that dont need scheduling
    pcomA = filter_courses(labs['pcomA'], lab_sched, c_hours, day)
    pcomB = filter_courses(labs['pcomB'], lab_sched, c_hours, day)
    bcomA = filter_courses(labs['bcomA'], lab_sched, c_hours, day)
    bcomB = filter_courses(labs['bcomB'], lab_sched, c_hours, day)
    
    
    if len(pcomA) > 0:
        lab_sched = add_lab(pcomA[0], pcomA_cohorts, c_hours, pcomA_strs,
                            lec_sched, lab_sched, week, day, "PCOM")
    if len(pcomB) > 0:
        lab_sched = add_lab(pcomB[0], pcomB_cohorts, c_hours,  pcomB_strs,
                            lec_sched, lab_sched, week, day, "PCOM")
    if len(bcomA) > 0:
        lab_sched = add_lab(bcomA[0], bcomA_cohorts, c_hours, bcomA_strs,
                            lec_sched, lab_sched, week, day, "BCOM")
    if len(bcomB) > 0:
        lab_sched = add_lab(bcomB[0], bcomB_cohorts, c_hours,  bcomB_strs,
                            lec_sched, lab_sched, week, day, "BCOM")
                
    return lab_sched
   

def make_prgm_lab_sched(lectures: Dict[str, List[Course]],
                        labs: Dict[str, List[Course]],
                        cohorts: Dict[str, List[str]], c_hours: Dict[str, int],
                        lec_sched: pd.DataFrame, lab_sched: pd.DataFrame, 
                        date_ints: Dict[str, int]) -> pd.DataFrame:
    '''
    Takes the existing lecture schedule and information on what labs to schedule,
    and tries to schedule each lab at a time when it won't conflict with lectures
    '''
    day  = date_ints['day']
    week = date_ints['week']
    
    # lists of lecture & lab IDs used to ensure no scheduling conflicts occur
    pmA_strs  = [c.ID for c in lectures['pmA']]  + [c.ID for c in labs['pmA']]
    pmB_strs  = [c.ID for c in lectures['pmB']]  + [c.ID for c in labs['pmB']]
    baA_strs  = [c.ID for c in lectures['baA']]  + [c.ID for c in labs['baA']]
    baB_strs  = [c.ID for c in lectures['baB']]  + [c.ID for c in labs['baB']]
    glmA_strs = [c.ID for c in lectures['glmA']] + [c.ID for c in labs['glmA']]
    glmB_strs = [c.ID for c in lectures['glmB']] + [c.ID for c in labs['glmB']]
    dxdA_strs = [c.ID for c in lectures['dxdA']] + [c.ID for c in labs['dxdA']]
    dxdB_strs = [c.ID for c in lectures['dxdB']] + [c.ID for c in labs['dxdB']]
    bkA_strs  = [c.ID for c in lectures['bkA']]  + [c.ID for c in labs['bkA']]
    bkB_strs  = [c.ID for c in lectures['bkB']]  + [c.ID for c in labs['bkB']]
    fsA_strs  = [c.ID for c in lectures['fsA']]  + [c.ID for c in labs['fsA']]
    fsB_strs  = [c.ID for c in lectures['fsB']]  + [c.ID for c in labs['fsB']]

    # parse cohorts for each program/term
    pmA_cohorts  = cohorts['pmA']
    pmB_cohorts  = cohorts['pmB']
    baA_cohorts  = cohorts['baA']
    baB_cohorts  = cohorts['baB']
    glmA_cohorts = cohorts['glmA']
    glmB_cohorts = cohorts['glmB']
    dxdA_cohorts = cohorts['dxdA']
    dxdB_cohorts = cohorts['dxdB']
    bkA_cohorts  = cohorts['bkA']
    bkB_cohorts  = cohorts['bkB']
    fsA_cohorts  = cohorts['fsA']
    fsB_cohorts  = cohorts['fsB']

    # filter out courses that dont need scheduling
    pmA  = filter_courses(labs['pmA'],  lab_sched, c_hours, day)
    pmB  = filter_courses(labs['pmB'],  lab_sched, c_hours, day)
    baA  = filter_courses(labs['baA'],  lab_sched, c_hours, day)
    baB  = filter_courses(labs['baB'],  lab_sched, c_hours, day)
    glmA = filter_courses(labs['glmA'], lab_sched, c_hours, day)
    glmB = filter_courses(labs['glmB'], lab_sched, c_hours, day)
    dxdA = filter_courses(labs['dxdA'], lab_sched, c_hours, day)
    dxdB = filter_courses(labs['dxdB'], lab_sched, c_hours, day)
    bkA  = filter_courses(labs['bkA'],  lab_sched, c_hours, day)
    bkB  = filter_courses(labs['bkB'],  lab_sched, c_hours, day)
    fsA  = filter_courses(labs['fsA'],  lab_sched, c_hours, day)
    fsB  = filter_courses(labs['fsB'],  lab_sched, c_hours, day)
    
    if len(fsB) > 0:
        lab_sched = add_lab(fsB[0], fsB_cohorts, c_hours, fsB_strs,
                            lec_sched, lab_sched, week, day, "FS", True)
    if len(fsA) > 0:
        lab_sched = add_lab(fsA[0], fsA_cohorts, c_hours, fsA_strs,
                            lec_sched, lab_sched, week, day, "FS", True)
    if len(dxdB) > 0:
        lab_sched = add_lab(dxdB[0], dxdB_cohorts, c_hours, dxdB_strs,
                            lec_sched, lab_sched, week, day, "DXD")
    if len(bkB) > 0:
        lab_sched = add_lab(bkB[0], bkB_cohorts, c_hours, bkB_strs,
                            lec_sched, lab_sched, week, day, "BK")
    if len(pmA) > 0:
        lab_sched = add_lab(pmA[0], pmA_cohorts, c_hours,  pmA_strs,
                            lec_sched, lab_sched, week, day, "PM")
    if len(pmB) > 0:
        lab_sched = add_lab(pmB[0], pmB_cohorts, c_hours,  pmB_strs,
                            lec_sched, lab_sched, week, day, "PM")
    if len(baA) > 0:
        lab_sched = add_lab(baA[0], baA_cohorts, c_hours,  baA_strs,
                            lec_sched, lab_sched, week, day, "BA")
    if len(baB) > 0:
        lab_sched = add_lab(baB[0], baB_cohorts, c_hours,  baB_strs,
                            lec_sched, lab_sched, week, day, "BA")
    if len(glmA) > 0:
        lab_sched = add_lab(glmA[0], glmA_cohorts, c_hours, glmA_strs,
                            lec_sched, lab_sched, week, day, "GLM")
    if len(glmB) > 0:
        lab_sched = add_lab(glmB[0], glmB_cohorts, c_hours, glmB_strs,
                            lec_sched, lab_sched, week, day, "GLM")
    if len(dxdA) > 0:
        lab_sched = add_lab(dxdA[0], dxdA_cohorts, c_hours, dxdA_strs,
                            lec_sched, lab_sched, week, day, "DXD")
    if len(bkA) > 0:
        lab_sched = add_lab(bkA[0], bkA_cohorts, c_hours,  bkA_strs,
                            lec_sched, lab_sched, week, day, "BK")
    
        
    return lab_sched
         

def add_lab(lab: Course, cohorts: List[str], course_hours: Dict[str, int], 
            invalid: List[str], lec_sched: pd.DataFrame, lab_sched: pd.DataFrame, 
            week: int, day: int, prgm_str: str, is_fs: bool = False) -> pd.DataFrame:
    '''
    Checks when/if a given lab can be scheduled without conflicting with any 
    lectures, what order the lab cohorts should be scheduled. Returns an updated 
    schedule if the lab can be added, otherwise this returns an unaltered schedule
    '''
    blocks = get_time_slot_count(lab, course_hours)

    # dict mapping room names to available start times (indexes, not strs)
    open_times = get_available_times(lab_sched, blocks, len(cohorts), is_fs)
    
    if (open_times == {}):
        return lab_sched

    cohort_times = {}
    for room, times in open_times.items():
        # for each start time, find out which cohorts can be scheduled
        for start in times:
            end = start + blocks
            valid = get_valid_cohorts(
                invalid, start, end, cohorts, room, lec_sched.join(lab_sched)
            )
            # only add cohorts we havent seen yet
            for cohort_ID in valid.difference(set(cohort_times.keys())):
                # only add rooms & start times that have not been assigned yet
                if (room, start) not in cohort_times.values():
                    cohort_times[cohort_ID] = (room, start)
                    
    # not all cohorts can fit in the schedule
    if len(cohort_times.keys()) < len(cohorts):
        return lab_sched
        
    for ID, time_slot in cohort_times.items():

        room, start = time_slot
        room_index  = lab_sched.columns.get_loc(room)
        start_label = list(lab_sched.index)[start]
        full_cohort = f"{prgm_str}0{lab.term}{ID}"

        # store the scheduled course's info in a lecture object
        add_lecture_to_db(
            Lecture(
                lab.ID, lab.title, lab.termHours, lab.term, lab.duration, 
                lab.isCore, lab.isOnline, lab.hasLab, lab.preReqs, full_cohort, 
                room, week, day, start_label   
            )
        )

        course_strs = [lab.ID + '-' + ID] * blocks
        lab_sched.iloc[start:(start+blocks), room_index] = course_strs

    return lab_sched

#============================== ONLINE SCHEDULING ==============================
def make_core_online_sched(lectures: Dict[str, List[Course]],
                           labs: Dict[str, List[Course]],
                           online: Dict[str, List[Course]], c_hours: Dict[str, int], 
                           curr_sched: pd.DataFrame, onl_sched: pd.DataFrame,
                           date_ints: Dict[str, int]) -> pd.DataFrame:
    '''
    Takes the existing lecture schedule and information on what online courses
    to schedule, and checks at each row (time) if an online course can be 
    scheduled without any conflicts. Courses also cant be within 1.5 hours 
    of in-person courses in the same program/term
    '''
    day  = date_ints['day']
    week = date_ints['week']
    
    # lists of course IDs used to ensure no scheduling conflicts occur
    pcomA_strs = [c.ID for c in lectures['pcomA']]+[c.ID for c in labs['pcomA']] 
    pcomB_strs = [c.ID for c in lectures['pcomB']]+[c.ID for c in labs['pcomB']]
    bcomA_strs = [c.ID for c in lectures['bcomA']]+[c.ID for c in labs['bcomA']]
    bcomB_strs = [c.ID for c in lectures['bcomB']]+[c.ID for c in labs['bcomB']]

    # filter out courses that dont need scheduling
    pcomA = filter_courses(online['pcomA'], onl_sched, c_hours, day)
    pcomB = filter_courses(online['pcomB'], onl_sched, c_hours, day)
    bcomA = filter_courses(online['bcomA'], onl_sched, c_hours, day)
    bcomB = filter_courses(online['bcomB'], onl_sched, c_hours, day)
    
    if len(pcomA) > 0:
            onl_sched = add_onl(pcomA[0], c_hours, pcomA_strs, curr_sched,
                                onl_sched, week, day, "PCOM")
    if len(pcomB) > 0:
            onl_sched = add_onl(pcomB[0], c_hours, pcomB_strs,  curr_sched,
                                onl_sched, week, day, "PCOM")
    if len(bcomA) > 0:
            onl_sched = add_onl(bcomA[0], c_hours, bcomA_strs,  curr_sched,
                                onl_sched, week, day, "BCOM")
    if len(bcomB) > 0:
            onl_sched = add_onl(bcomB[0], c_hours, bcomB_strs,  curr_sched,
                                onl_sched, week, day, "BCOM")

    return onl_sched


def make_prgm_online_sched(lectures: Dict[str, List[Course]],
                           labs: Dict[str, List[Course]],
                           online: Dict[str, List[Course]], c_hours: Dict[str, int], 
                           curr_sched: pd.DataFrame, onl_sched: pd.DataFrame,
                           date_ints: Dict[str, int]) -> pd.DataFrame:
    '''
    Takes the existing lecture schedule and information on what online courses
    to schedule, and checks at each row (time) if an online course can be 
    scheduled without any conflicts. Courses also cant be within 1.5 hours 
    of in-person courses in the same program/term
    '''
    day  = date_ints['day']
    week = date_ints['week']
    
    # lists of lecture & lab IDs used to ensure no scheduling conflicts occur
    pmA_strs  = [c.ID for c in lectures['pmA']]  + [c.ID for c in labs['pmA']]
    pmB_strs  = [c.ID for c in lectures['pmB']]  + [c.ID for c in labs['pmB']]
    baA_strs  = [c.ID for c in lectures['baA']]  + [c.ID for c in labs['baA']]
    baB_strs  = [c.ID for c in lectures['baB']]  + [c.ID for c in labs['baB']]
    glmA_strs = [c.ID for c in lectures['glmA']] + [c.ID for c in labs['glmA']]
    glmB_strs = [c.ID for c in lectures['glmB']] + [c.ID for c in labs['glmB']]
    dxdA_strs = [c.ID for c in lectures['dxdA']] + [c.ID for c in labs['dxdA']]
    dxdB_strs = [c.ID for c in lectures['dxdB']] + [c.ID for c in labs['dxdB']]
    bkA_strs  = [c.ID for c in lectures['bkA']]  + [c.ID for c in labs['bkA']]
    bkB_strs  = [c.ID for c in lectures['bkB']]  + [c.ID for c in labs['bkB']]
    fsA_strs  = [c.ID for c in lectures['fsA']]  + [c.ID for c in labs['fsA']]
    fsB_strs  = [c.ID for c in lectures['fsB']]  + [c.ID for c in labs['fsB']]

    # filter out courses that dont need scheduling
    pmA  = filter_courses(online['pmA'],  onl_sched, c_hours, day)
    pmB  = filter_courses(online['pmB'],  onl_sched, c_hours, day)
    baA  = filter_courses(online['baA'],  onl_sched, c_hours, day)
    baB  = filter_courses(online['baB'],  onl_sched, c_hours, day)
    glmA = filter_courses(online['glmA'], onl_sched, c_hours, day)
    glmB = filter_courses(online['glmB'], onl_sched, c_hours, day)
    dxdA = filter_courses(online['dxdA'], onl_sched, c_hours, day)
    dxdB = filter_courses(online['dxdB'], onl_sched, c_hours, day)
    bkA  = filter_courses(online['bkA'],  onl_sched, c_hours, day)
    bkB  = filter_courses(online['bkB'],  onl_sched, c_hours, day)
    fsA  = filter_courses(online['fsA'],  onl_sched, c_hours, day)
    fsB  = filter_courses(online['fsB'],  onl_sched, c_hours, day)
    
    if len(pmA) > 0:
            onl_sched = add_onl(pmA[0], c_hours, pmA_strs,curr_sched,
                                onl_sched, week, day, "PM")
    if len(pmB) > 0:
            onl_sched = add_onl(pmB[0], c_hours, pmB_strs, curr_sched,
                                onl_sched, week, day, "PM")
    if len(baA) > 0:
            onl_sched = add_onl(baA[0], c_hours, baA_strs, curr_sched,
                                onl_sched, week, day, "BA")
    if len(baB) > 0:
            onl_sched = add_onl(baB[0], c_hours, baB_strs, curr_sched,
                                onl_sched, week, day, "BA")
    if len(glmA) > 0:
            onl_sched = add_onl(glmA[0], c_hours, glmA_strs, curr_sched,
                                onl_sched, week, day, "GLM")
    if len(glmB) > 0:
            onl_sched = add_onl(glmB[0], c_hours, glmB_strs, curr_sched,
                                onl_sched, week, day, "GLM")
    if len(dxdA) > 0:
            onl_sched = add_onl(dxdA[0], c_hours, dxdA_strs, curr_sched,
                                onl_sched, week, day, "DXD")
    if len(dxdB) > 0:
            onl_sched = add_onl(dxdB[0], c_hours, dxdB_strs, curr_sched,
                                onl_sched, week, day, "DXD")
    if len(bkA) > 0:
            onl_sched = add_onl(bkA[0], c_hours, bkA_strs, curr_sched,
                                onl_sched, week, day, "BK")
    if len(bkB) > 0:
            onl_sched = add_onl(bkB[0], c_hours, bkB_strs, curr_sched,
                                onl_sched, week, day, "BK")
    if len(fsA) > 0:
            onl_sched = add_onl(fsA[0], c_hours, fsA_strs, curr_sched,
                                onl_sched, week, day, "FS", True)
    if len(fsB) > 0:
            onl_sched = add_onl(fsB[0], c_hours, fsB_strs, curr_sched,
                                onl_sched, week, day, "FS", True)
    return onl_sched


def add_onl(online: Course, course_hours: Dict[str, int], 
            invalid: List[str], curr_sched: pd.DataFrame, onl_sched: pd.DataFrame, 
            week: int, day: int, prgm_str: str, is_fs: bool = False) -> pd.DataFrame:
    '''
    Checks the current schedule to see if an online course can be 
    scheduled without any conflicts. Courses cannot be within 1.5 hours of an
    in-person class in the same program/term. Cohorts are not required since 
    online courses are not restricted by classroom capacity
    '''

    blocks = get_time_slot_count(online, course_hours)

    # dict mapping room names to available start times (indexes, not strs)
    open_times = get_available_times(onl_sched, blocks, 1, is_fs)
    
    if (open_times == {}):
        return onl_sched
    
    # check if each timeslot is not within 1.5 hours of a lecture 
    for times in open_times.values():
        for start in times:
            end = start + blocks
            
            if can_sched_online(curr_sched, invalid, start, end):
                
                cohort = f"{prgm_str}0{online.term}01"
                room   = list(onl_sched.columns)[0]
                s_time = onl_sched.index[start]
                
                add_lecture_to_db(Lecture(
                    online.ID, online.title, online.termHours, online.term, 
                    online.duration, online.isCore, online.isOnline, online.hasLab, 
                    online.preReqs, cohort, room, week, day, s_time 
                ))
                
                onl_sched.iloc[start:end, 0] = ([online.ID] * blocks)
                break
    
    return onl_sched

def can_sched_online(curr_sched: pd.DataFrame, invalid_courses: List[str], 
                     start: int, end: int) -> bool:
    '''
    Compares a possible start & end time for an online course and returns a 
    boolean indicating if the course can be scheduled at that time.
    '''

    for room in curr_sched.columns:
        for idx in range(max(start-3, 0), min(end+2, len(curr_sched.index))):
            if (curr_sched.iloc[idx][room])[:-3] in invalid_courses:
                return False
    return True
#===============================================================================

def make_core_day_sched(data: Dict[str, Union[List[Course], List[str]]],
                        c_hours: Dict[str, int], date_ints: Dict[str, int],
                        prev_scheds: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    '''
    Creates a lecture, lab, and online course schedule for a monday/wednesday
    Returns a single DataFrame of the schedules joined together
    '''
    lecs = data["lectures"]
    labs = data["labs"]
    onls = data["online"]
    cohorts = data["cohorts"]
    
    prev_lecs = prev_scheds["lecture"]
    prev_labs = prev_scheds["lab"]
    prev_onls = prev_scheds["online"]
    
    lec_sched = make_core_lecture_sched(lecs, cohorts, 
                                        c_hours, prev_lecs, date_ints)
    
    lab_sched = make_core_lab_sched(lecs, labs, cohorts, c_hours, 
                                    lec_sched, prev_labs, date_ints)
    
    joined_sched = lec_sched.join((lab_sched))
    
    onl_sched = make_core_online_sched(lecs, labs, onls, c_hours, 
                                       joined_sched, prev_onls, date_ints)
    
    return joined_sched.join(onl_sched)

def make_prgm_day_sched(data: Dict[str, Union[List[Course], List[str]]],
                        c_hours: Dict[str, int], date_ints: Dict[str, int],
                        prev_scheds: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    '''
    Creates a lecture, lab, and online course schedule for a monday/wednesday
    Returns a single DataFrame of the schedules joined together
    '''
    lecs = data["lectures"]
    labs = data["labs"]
    onls = data["online"]
    cohorts = data["cohorts"]
    
    prev_lecs = prev_scheds["lecture"]
    prev_labs = prev_scheds["lab"]
    prev_onls = prev_scheds["online"]
    
    lec_sched = make_prgm_lecture_sched(lecs, cohorts, 
                                        c_hours, prev_lecs, date_ints)
    
    lab_sched = make_prgm_lab_sched(lecs, labs, cohorts, c_hours, 
                                    lec_sched, prev_labs, date_ints)
    
    joined_sched = lec_sched.join((lab_sched))
    
    onl_sched = make_prgm_online_sched(lecs, labs, onls, c_hours, 
                                       joined_sched, prev_onls, date_ints)
    
    return joined_sched.join(onl_sched)

def getHolidaysMonWed(fallYear):
    '''
    Pass the year of the fall term, and will make a list of holidays that land on 
    mondays and wednesdays in the 3 terms. Returns list of dt objects
    '''
    holidayList = []
    nextYear = fallYear + 1
    FallmonthList = [9,10,11]
    TARD = dt.date(fallYear,9,30)
    nextMonthList = [1,2,3,4, 5,6,7,8]
    
    #FALL:
    for ptr in holidays.Canada(years = fallYear).items():
        if ( ptr[0].month in FallmonthList):
            if("Observed" not in ptr[1] and (ptr[0].weekday() == 0 or ptr[0].weekday() == 2)):
                holidayList.append(ptr[0])
    if(TARD.weekday()== 0 or TARD.weekday() == 2):
        holidayList.insert(1,TARD)
    #WIN/SPRING OF NEXT YEAR
    for ptr in holidays.Canada(years = nextYear).items():
        if ( ptr[0].month in nextMonthList and (ptr[0].weekday() == 0 or ptr[0].weekday() == 2)):
            if("Observed" not in ptr[1] and 'New Year' not in ptr[1]):
                holidayList.append(ptr[0])
    return holidayList  

def getHolidaysTuesThurs(fallYear):
    '''
    Pass the year of the fall term, and will make a list of holidays that land on 
    Tuesdays and Thursdays in the 3 terms. Returns list of dt objects
    '''
    holidayList = []
    nextYear = fallYear + 1
    FallmonthList = [9,10,11]
    TARD = dt.date(fallYear,9,30)
    nextMonthList = [1,2,3,4, 5,6,7,8]

    #FALL:
    for ptr in holidays.Canada(years = fallYear).items():
        if ( ptr[0].month in FallmonthList):
            if("Observed" not in ptr[1] and (ptr[0].weekday() == 1 or ptr[0].weekday() == 3)):
                holidayList.append(ptr[0])
    if(TARD.weekday()== 1 or TARD.weekday() == 3):
        holidayList.insert(1,TARD)

    #WIN/SPRING OF NEXT YEAR
    for ptr in holidays.Canada(years = nextYear).items():
        if ( ptr[0].month in nextMonthList and (ptr[0].weekday() == 1 or ptr[0].weekday() == 3)):
            if("Observed" not in ptr[1] and 'New Year' not in ptr[1]):
                holidayList.append(ptr[0])  
    return holidayList  


'''reading week in fall - happens after week with remb.day or week after if it lands on a weekend'''
'''reading week in wint - happens after week with fam day in feb or week after if it lands on a weekend '''
'''No reading break in sp/su'''

def getFallStartDay(year):
    '''Returns datetime object of the first day of fall term in passed year
    first Wednesday of September '''
    sept1 = dt.date(year,9,1)
    offset = 2-sept1.weekday() #weekday = 2 means wednesday
    if offset < 0:
        offset+=7
    return sept1+dt.timedelta(offset)

def getWinterStartDay(year):
    '''Returns datetime object of the first day of winter term in passed year
    first Wednesday of Janurary'''
    jan1 = dt.date(year,1,1)
    offset = 2-jan1.weekday() #weekday = 2 means wednesday
    if offset < 0:
        offset+=7
    return jan1+dt.timedelta(offset)
def getSpringStartDay(year):
    '''Returns datetime object of the first day of spring term in passed year
    May 1st if lands on a weekday, otherwise the first Monday of May'''
    startday = dt.date(year,5,1) #may 1
    if (startday.weekday() == 5):
        startday = dt.date(year,5,3)
    if(startday.weekday() == 6):
        startday = dt.date(year,5,2)
    return startday

