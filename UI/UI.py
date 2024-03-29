# main python file - 395 team 8

#Imports
import os
from PyQt5 import QtGui
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
import random
import fill_data

from UI import widgets as w
from UI import remove_colours as rc
import help_funcs

from database.database import *
import imports.fillClassrooms

import imports.schedulers.core_scheduler
import imports.schedulers.program_scheduler
import imports.classes.classrooms
import datetime
import copy

SEM = {"Fall":1, "Winter":2, "Spring / Summer":3}
# Quickly match times to slots when displaying schedule.
TIMES = {"08:00":0, "08:30":1,
         "09:00":2, "09:30":3,
         "10:00":4,"10:30":5,
         "11:00":6,"11:30":7,
         "12:00":8,"12:30":9,
         "13:00":10,"13:30":11,
         "14:00":12,"14:30":13,
         "15:00":14,"15:30":15,
         "16:00":16,"16:30":17,
         "17:00":18, "17:30":19,
         "18:00":20, "18:30":21,
         "19:00":22, "19:30":23,
         "20:00":24, "20:30":25}

LEFT_MAX_WIDTH = 450

# Dictionaries where key = week number, value =list of (2) lists for each day that week
(CORE_SCHEDULE, PROG_SCHEDULE) = ({}, {})
(CORE_SCHEDULE_COHORTS, PROG_SCHEDULE_COHORTS) = ({}, {})

(CORE_END_DATES, PROG_END_DATES) = ({}, {})
(CORE_HOLIDAYS, PROG_HOLIDAYS) = ([], [])

WEEK_DISPLAY_DATE = {}
ROOM = ""
global PROG_LABELS
(WEEK, CORE_DAY, PROG_DAY) = (1, 1, 1)
COLOUR_INDEX = -1
COURSE_COLOUR = {}
COHORT_COURSE_COLOUR = {}

WEEK_COHORTS = 1
COHORT_CHOSEN = ""
COHORT_COURSE_TO_ROOM = {}

CREATE_SCHEDULE_CLICKED = 0

BG_COLOURS = QtGui.QColor.colorNames()

class UI(QMainWindow):

    def __init__(self):
        super().__init__()

        global BG_COLOURS
        BG_COLOURS = rc.remove_colours(BG_COLOURS)

        self.setWindowTitle("Scheduler")
        self.setStyleSheet("background-color: #6f2937") 

        # Create references for things that can change - filepaths, charts etc.\
        # Can add more as needed
        self.file_path = ""
        self.file_label = w.label(w.snow_reg, "No File Chosen")
        self.select_room = w.drop_down(w.glass, self.room_selector_show_schedule)

        self.week_label = w.label(w.snow_header1, "", 16)
        self.week_label.setAlignment(Qt.AlignCenter)

        self.cohort_week_label = w.label(w.snow_header1, "", 16)
        self.cohort_week_label.setAlignment(Qt.AlignCenter)

        self.pick_semester = w.drop_down(w.glass)
        self.pick_semester.addItems(list(SEM.keys()))

        self.export_check = QCheckBox()
        self.export_check.setStyleSheet("color: white")
        self.export_check

        # Options
        self.class_id = QLineEdit()
        self.class_lab = QButtonGroup()
        self.class_capacity = w.spin_box(w.glass)
        self.classroom_list = w.drop_down(w.glass)
        self.courses = w.drop_down(w.glass)
        self.courses_edit_new = QButtonGroup()
        self.course_id = QLineEdit()
        self.course_term_hours = w.spin_box(w.glass)
        self.course_term = w.spin_box(w.glass)
        self.course_term.setStyleSheet(w.glass.specs)
        self.course_duration = w.spin_box(w.glass)
        self.course_duration.setStyleSheet(w.glass.specs) 
        self.course_core = QCheckBox()
        self.course_online = QCheckBox()
        self.course_lab = QCheckBox()
        self.course_program = w.drop_down(w.glass)


        self.course_pre_req_selector = w.drop_down(w.glass)
        self.course_pre_reqs = []
        self.course_pre_reqs_label = w.label(w.snow_reg, "")

        '''
        Creating tables for each tab
        and giving proper settings (i.e. un-editable, resize to width etc)
        '''
        self.main_table = QTableWidget()

        self.main_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.main_table.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # Make table un-editable / un-targettable
        self.main_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.main_table.setSelectionMode(QAbstractItemView.NoSelection)
        self.main_table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.main_table.setShowGrid(False)
        
        '''
        Table for the Cohort filter schedule
        + Combobox for picking cohort
        '''
        self.cohort_table = QTableWidget()
        self.cohort_tab_combo = w.drop_down(w.glass, self.cohort_selector_show_schedule)

        self.cohort_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.cohort_table.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # Make table un-editable / un-targettable
        self.cohort_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.cohort_table.setSelectionMode(QAbstractItemView.NoSelection)
        self.cohort_table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.cohort_table.setShowGrid(False)

        '''
        Layouts containing the term inputs
        for easier iteration when reading inputs
        '''
        self.term_1_inputs = QVBoxLayout()
        self.term_2_inputs = QVBoxLayout()
        self.term_3_inputs = QVBoxLayout()

        central_widget = QWidget()  # Must have central widget set
        central_widget.setLayout(self.create_hlayout())

        self.setCentralWidget(central_widget)

        self.show()

    # Creates all items for central widget
    def create_hlayout(self):
        hbox = QHBoxLayout(self)
        right_box = self.create_tabs()
        left_vbox = self.create_leftlayout()

        # Add layouts to overall area
        hbox.addWidget(left_vbox)
        hbox.addWidget(right_box)

        hbox.setAlignment(left_vbox, Qt.AlignTop)

        return hbox

    # Creates the top hbox where most information will be displayed
    def create_tabs(self):

        # Create tabs
        tabs = QTabWidget()
        tabs.setStyleSheet(w.glass.specs)

        (tab1, tab2, tab3, tab4) = (QWidget(), QWidget(), QWidget(), QWidget())

        t_list = [tab1, tab2, tab3, tab4]
        styles = [w.brick.specs, w.brick.specs, w.brick.specs, w.paper.specs]
        t_titles = ["Classroom Schedule", "Cohort Schedule", "Options", "Instructions"]
        t_funcs = [self.make_main_tab(), self.make_cohort_sched_tab(), 
                   self.make_options_tab(), self.make_readme_tab()]

        for i in range(len(t_list)):
            t_list[i].setStyleSheet(styles[i])
            tabs.addTab(t_list[i], t_titles[i])
            t_list[i].setLayout(t_funcs[i])

        self.create_schedule_base(0)
        self.create_cohorts_schedule_base()

        return tabs


    def make_main_tab(self):
        main_table_box = QVBoxLayout(self)
        week_choose = QHBoxLayout(self)

        #Left/Right Navigation Arrows
        left =  w.push_button(w.coal,"←", self.back_week)
        right = w.push_button(w.coal,"→", self.forward_week)

        week_choose.addWidget(left)
        week_choose.addWidget(self.week_label)
        week_choose.addWidget(right)

        main_table_box.addLayout(week_choose)
        main_table_box.addWidget(self.main_table)

        return main_table_box
    

    def make_readme_tab(self):
        # Read me Doc
        read_me = QTextEdit()
        rdme_path = help_funcs.check_path('README.md')
        try:
            file_content = open(rdme_path).read()
        except:
            rdme_path = help_funcs.check_path('..\\README.md')
            file_content = open(rdme_path).read()
        read_me.setMarkdown(file_content)
        layout = QHBoxLayout()
        layout.addWidget(read_me)

        return layout

    def make_options_tab(self):
        vbox_overall = QVBoxLayout()

        vbox_overall.addLayout(self.update_classroom_section())
        vbox_overall.addWidget(w.create_horizontal_line())
        vbox_overall.addWidget(w.create_horizontal_line())

        vbox_overall.addLayout(self.update_course_section())
        vbox_overall.addWidget(w.create_horizontal_line())
        vbox_overall.addWidget(w.create_horizontal_line())

        #Reset database button
        reset_label = w.label(w.snow_header1, "Reset Database")

        reset_button = w.push_button(w.glass, "Reset to Default Settings", self.reset_db)
        reset_button.setFixedWidth(200)

        hbox = QVBoxLayout()
        hbox.setContentsMargins(20,40,0,50)
        hbox.addWidget(reset_label)
        hbox.addWidget(reset_button)

        vbox_overall.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.MinimumExpanding, QSizePolicy.Minimum))
        vbox_overall.addLayout(hbox)
        vbox_overall.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.MinimumExpanding, QSizePolicy.Minimum))

        return vbox_overall


    def reset_db(self):

        global CREATE_SCHEDULE_CLICKED, CORE_SCHEDULE, CORE_SCHEDULE_COHORTS, PROG_SCHEDULE_COHORTS, PROG_SCHEDULE
        answer = QMessageBox(QMessageBox.Warning, "Reset Database",
                                        "Are you sure you want to reset the database?\n\nAll data will be lost!",
                                        buttons=QMessageBox.Yes | QMessageBox.No)
        answer.setDefaultButton(QMessageBox.No)
        answer.setStyleSheet("color: black")
        clicked = answer.exec()

        if clicked == QMessageBox.Yes:
            try:
                os.remove("./database/database.db")
                fill_data.createDefaultDatabase()
                self.update_class_combos()
                self.update_course_combos()
                self.cohort_tab_combo.clear()
                self.reset_table()
                self.reset_table_cohort()
                CORE_SCHEDULE_COHORTS.clear()
                PROG_SCHEDULE_COHORTS.clear()
                CORE_SCHEDULE.clear()
                PROG_SCHEDULE.clear()
                CREATE_SCHEDULE_CLICKED = 0

            except:
                return
        else:
            return

    def update_classroom_section(self):
        room_add_layout = QHBoxLayout()
        room_add_layout.setSpacing(15)

        room_delete_layout = QHBoxLayout()
        room_delete_layout.setSpacing(15)

        # Classroom section
        vbox_class = QVBoxLayout()
        vbox_class.setContentsMargins(20,0,0,0)
        class_section_text = w.label(w.snow_header1, "Room Options")

        #Header
        vbox_class.addWidget(class_section_text)
        vbox_class.addWidget(w.create_horizontal_line())

        #Adding a Room Subheader
        class_add_text = w.label(w.snow_header2, "Add New Room")
        vbox_class.addWidget(class_add_text)

        # Class ID section
        self.class_id.setPlaceholderText("Classroom Name")
        self.class_id.setStyleSheet("color: #fefdea")
        class_id_box = QHBoxLayout()
        class_id_label = w.label(w.snow_reg, "Room ID")
        class_id_box.addWidget(class_id_label)
        #self.class_id.setMaximumWidth(100)
        class_id_box.addWidget(self.class_id)

        # Class Capacity Section
        class_capacity_box = QHBoxLayout()
        room_cap_label = w.label(w.snow_reg, "Room Capacity")
        class_capacity_box.addWidget(room_cap_label)
        self.class_capacity.setValue(10)
        self.class_capacity.setMinimum(10)
        self.class_capacity.setMaximum(50)
        self.class_capacity.setMaximumWidth(100)
        class_capacity_box.addWidget(self.class_capacity)

        # Class Lab/Lecture section
        class_lab_bool = QHBoxLayout()
        b1 = QRadioButton("Lab")
        b1.setStyleSheet("color: #fefdea")
        b1.setChecked(True)
        b2 = QRadioButton("Lecture")
        b2.setStyleSheet("color: #fefdea")
        self.class_lab.addButton(b1)
        self.class_lab.addButton(b2)

        lab_bool_label = w.label(w.snow_reg, "Room Type")
        class_lab_bool.addWidget(lab_bool_label)
        class_lab_bool.addWidget(b1)
        class_lab_bool.addWidget(b2)

        # Create add button
        class_btn = w.push_button(w.glass, "Add", self.add_edit_classroom)
        class_btn.setMaximumWidth(100)

        # Put adding functions into the layout
        room_add_layout.addLayout(class_id_box)
        room_add_layout.addWidget(w.create_vertical_line())
        room_add_layout.addLayout(class_capacity_box)
        room_add_layout.addWidget(w.create_vertical_line())
        room_add_layout.addLayout(class_lab_bool)
        room_add_layout.addWidget(class_btn)
        
        #Deleting a Room

        #Delete Room Header
        class_delete_text = w.label(w.snow_header2, "Delete Room")

        # Create remove button
        remove_btn = w.push_button(w.glass, "Remove", self.remove_classroom)
        remove_btn.setMaximumWidth(150)

        remove_section = QHBoxLayout()
        remove_section.addWidget(self.classroom_list)
        remove_section.addWidget(remove_btn)

        # Put deleting functions into layout
        room_delete_layout.addLayout(remove_section)
        room_delete_layout.addSpacerItem(QSpacerItem(10, 10, QSizePolicy.MinimumExpanding, QSizePolicy.Minimum))

        #Add sublayouts to big vbox_class object
        vbox_class.addLayout(room_add_layout)
        vbox_class.addWidget(class_delete_text)
        vbox_class.addLayout(room_delete_layout)
        vbox_class.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.Minimum, QSizePolicy.Minimum))
        
        return vbox_class


    def update_course_section(self):
        vbox_course = QVBoxLayout()
        vbox_course.setContentsMargins(20,0,0,0)

        course_section_text = w.label(w.snow_header1, "Course Options")
        
        vbox_course.addWidget(course_section_text)
        vbox_course.addWidget(w.create_horizontal_line())
        vbox_course.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.Minimum, QSizePolicy.Minimum))

        db = help_funcs.check_path("database\database.db")  # database.db file path
        connection = create_connection(db)
        db_courses = readCourseItem(connection, "%")
        close_connection(connection)

        course_name_list = []

        for each_course in range(len(db_courses)):
            course_name_list.append(db_courses[each_course][0])

        self.courses.addItems(course_name_list)
        self.courses.setMinimumWidth(100)
        self.courses.setMaximumWidth(100)
        self.course_pre_req_selector.addItems(course_name_list)

        # Creating inputs
        hbox = QHBoxLayout()

        new_or_edit = QVBoxLayout()
        b1 = QRadioButton("New Course")
        b1.setStyleSheet("color: #fefdea")
        b1.setChecked(True)
        b2 = QRadioButton("Edit Course")
        b2.setStyleSheet("color: #fefdea")
        self.courses_edit_new.addButton(b1)
        self.courses_edit_new.addButton(b2)

        new_or_edit.addWidget(b1)
        new_or_edit.addWidget(b2)
        #-----------------------------

        hbox_program = QHBoxLayout()
        prog_label = w.label(w.snow_reg, "Program: ")
        hbox_program.addWidget(prog_label)
        self.course_program.addItems(["PCOM", "BCOM", "PM", "BA", "GLM", "FS", "DXD", "BK"])
        hbox_program.addWidget(self.course_program)

        # -----------------------------

        course_id_sec = QHBoxLayout()
        course_name_label = w.label(w.snow_reg, "Course Name")
        course_id_sec.addWidget(course_name_label)
        self.courses_edit_new.buttonClicked.connect(self.show_hide_course)
        self.course_id.setStyleSheet("color: #fefdea")
        self.course_id.setMinimumWidth(200)
        self.course_id.setMaximumWidth(200)
        course_id_sec.addWidget(self.courses)
        course_id_sec.addWidget(self.course_id)
        self.courses.hide()

        # ---------------------------

        vbox_spin_boxes = QVBoxLayout()
        hbox_hours = QHBoxLayout()
        hours_label = w.label(w.snow_reg, "Course Hours: ")
        hbox_hours.addWidget(hours_label)
        hbox_hours.addWidget(self.course_term_hours)

        hbox_term = QHBoxLayout()
        term_label = w.label(w.snow_reg, "Term: ")
        hbox_term.addWidget(term_label)
        hbox_term.addWidget(self.course_term)

        hbox_duration = QHBoxLayout()
        duration_label = w.label(w.snow_reg, "Duration: ")
        hbox_duration.addWidget(duration_label)
        hbox_duration.addWidget(self.course_duration)

        vbox_spin_boxes.addLayout(hbox_hours)
        vbox_spin_boxes.addLayout(hbox_term)
        vbox_spin_boxes.addLayout(hbox_duration)

        #--------------------------

        hbox_online_lab = QHBoxLayout()

        core_hbox = QHBoxLayout()
        core_label = w.label(w.snow_reg, "Core Course: ")
        core_hbox.addWidget(core_label)
        core_hbox.addWidget(self.course_core)

        online_hbox = QHBoxLayout()
        online_label = w.label(w.snow_reg, "Online: ")
        online_hbox.addWidget(online_label)
        online_hbox.addWidget(self.course_online)

        hbox_lab = QHBoxLayout()
        lab_comp_label = w.label(w.snow_reg, "Lab Component: ")
        hbox_lab.addWidget(lab_comp_label)
        hbox_lab.addWidget(self.course_lab)

        hbox_online_lab.addLayout(core_hbox)
        hbox_online_lab.addLayout(online_hbox)
        hbox_online_lab.addLayout(hbox_lab)
        #--------------------------

        course_btn = w.push_button(w.glass, "Save Course", self.save_course)

        # --------------------------

        hbox_pre_reqs = QHBoxLayout()
        vbox_pre_reqs = QVBoxLayout()

        pre_req_label = w.label(w.snow_reg, "Chosen Pre-Reqs")
        hbox_pre_reqs.addWidget(pre_req_label)
        add_pre_req = w.push_button(w.glass, "Add Pre-Req", self.add_pre_req)
        rem_pre_req = w.push_button(w.glass, "Clear Pre-Reqs", self.clear_pre_reqs)


        hbox_pre_reqs.addWidget(self.course_pre_reqs_label)
        vbox_pre_reqs.addLayout(hbox_pre_reqs)
        vbox_pre_reqs.addWidget(self.course_pre_req_selector)
        vbox_pre_reqs.addWidget(add_pre_req)
        vbox_pre_reqs.addWidget(rem_pre_req)

        #-----------------------------
        vbox_course_specs1 = QVBoxLayout()
        vbox_course_specs1.addLayout(hbox_program)
        vbox_course_specs1.addLayout(course_id_sec)
        vbox_course_specs1.addLayout(hbox_online_lab)
        #-----------------------------

        hbox.setSpacing(15)
        hbox.addLayout(new_or_edit)
        hbox.addWidget(w.create_vertical_line())
        hbox.addLayout(vbox_course_specs1)
        hbox.addWidget(w.create_vertical_line())
        hbox.addLayout(vbox_spin_boxes)
        hbox.addWidget(w.create_vertical_line())
        hbox.addLayout(vbox_pre_reqs)
        hbox.addWidget(w.create_vertical_line())
        hbox.addWidget(course_btn)

        vbox_course.addLayout(hbox)

        return vbox_course

    '''
    The following 2 functions
    create the basic layout 
    for the tables
    cohorts and main
    '''
    def create_schedule_base(self, isLab):

        days = ["Monday", "Tuesday", "Wednesday", "Thursday"]
        times = []

        # Creation of all times from 8:00 AM to 5:00 PM to use as row headers

        first_time = datetime.datetime(year=2000, month=1, day=1, hour=8, minute=00)
        time_dif = datetime.timedelta(minutes=30)

        times.append(first_time.strftime("%I:%M %p"))
        new_time = first_time + time_dif
        if (isLab):
            for half_hour in range(25):
                times.append(new_time.strftime("%I:%M %p"))
                new_time = new_time + time_dif

        else:
            for half_hour in range(18):
                times.append(new_time.strftime("%I:%M %p"))
                new_time = new_time + time_dif

        self.main_table.setColumnCount(4)
        self.main_table.setHorizontalHeaderLabels(days)

        self.main_table.setRowCount(len(times))
        self.main_table.setVerticalHeaderLabels(times)

        # Fill with empty items to change background colours later
        self.reset_table()

    # Creates the bottom layout where most user interaction takes place
    def create_leftlayout(self):

        width_limit = QWidget()
        width_limit.setMaximumWidth(LEFT_MAX_WIDTH)

        vbox = QVBoxLayout(self)
        vbox.setSizeConstraint(QLayout.SetFixedSize) # Prevents left side from resizing

        #Big Scheduler label in top left corner
        title = w.label(w.snow_header1, "Scheduler", 40)
        title.setMaximumWidth(LEFT_MAX_WIDTH)

        #Students Per Term Label
        input_title = w.label(w.snow_header1, "Students per Term", 18)
        input_title.setMaximumWidth(LEFT_MAX_WIDTH)

        #Term & Classroom label
        select_title = w.label(w.snow_header1, "Term & Classroom", 18)
        select_title.setMaximumWidth(LEFT_MAX_WIDTH)

        #Big green Create Schedule Button
        create_sched = w.push_button(w.coal, "Create Schedule", self.create_schedule)
        create_sched.setFixedSize(280,60)

        # Export Checkbox
        export_label = w.label(w.snow_header1, "Export to File:", 14)
        self.export_check.setCheckState(True)
        self.export_check.setTristate(False)
        export_hbox = QHBoxLayout()
        export_hbox.addWidget(export_label)
        export_hbox.addWidget(self.export_check)


        # Read Current items in the Database
        self.update_class_combos()


        vbox.addWidget(title)
        vbox.addWidget(w.create_horizontal_line())
        vbox.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.Minimum, QSizePolicy.Minimum))
        vbox.addLayout(self.create_file_choose())
        vbox.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.Minimum, QSizePolicy.Minimum))
        vbox.addWidget(w.create_horizontal_line())
        vbox.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.Minimum, QSizePolicy.Minimum))
        vbox.addWidget(input_title)
        vbox.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.Minimum, QSizePolicy.Minimum))
        vbox.addLayout(self.legion_inputs())
        vbox.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.Minimum, QSizePolicy.Minimum))
        vbox.addWidget(w.create_horizontal_line())
        vbox.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.Minimum, QSizePolicy.Minimum))
        vbox.addWidget(select_title)
        vbox.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.Minimum, QSizePolicy.Minimum))
        vbox.addWidget(self.pick_semester)
        vbox.addWidget(self.select_room)
        vbox.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.Minimum, QSizePolicy.Minimum))
        vbox.addWidget(w.create_horizontal_line())
        vbox.addSpacerItem(QSpacerItem(15, 15, QSizePolicy.Minimum, QSizePolicy.Minimum))
        vbox.addWidget(create_sched)
        vbox.addLayout(export_hbox)
        

        width_limit.setLayout(vbox)

        return width_limit

    # create the file choose layout / widgets
    def create_file_choose(self):

        vbox = QVBoxLayout()

        #Title
        title = w.label(w.snow_header1, "Enrollment Import", 18)
        title.setStyleSheet("color: #fefdea")
        title.setMaximumWidth(LEFT_MAX_WIDTH)

        hbox_file_choose = QHBoxLayout(self)
        choose_input_button = w.push_button(w.glass, "Choose File", self.choose_file)
        choose_input_button.setMaximumWidth(100)

        self.file_label.setMaximumWidth(LEFT_MAX_WIDTH)

        hbox_file_choose.addWidget(choose_input_button)
        hbox_file_choose.addWidget(self.file_label)

        load_data = w.push_button(w.glass, "Load Data", self.load_student_numbers)

        create_template_btn = w.push_button(w.glass, "Create Template", self.create_template)

        vbox.addWidget(title)
        vbox.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.Minimum, QSizePolicy.Minimum))
        vbox.addLayout(hbox_file_choose)
        vbox.addWidget(load_data)
        vbox.addWidget(create_template_btn)

        return vbox

    # Makes layout for the term input section
    def legion_inputs(self):
        vbox_all = QVBoxLayout()

        hbox_inputs = QHBoxLayout()

        hbox_inputs.addLayout(self.program_labels())
        hbox_inputs.addLayout(self.create_term_inputs(self.term_1_inputs, "Term 1"))
        hbox_inputs.addLayout(self.create_term_inputs(self.term_2_inputs, "Term 2"))
        hbox_inputs.addLayout(self.create_term_inputs(self.term_3_inputs, "Term 3"))

        vbox_all.addLayout(hbox_inputs)

        #Load saved student numbers button
        load_saved_students = w.push_button(w.glass, 
                                            "Load Saved Student Numbers", 
                                            self.load_db_stu_num)

        vbox_all.addWidget(load_saved_students)


        return vbox_all


    def program_labels(self):
        global PROG_LABELS
        vbox_labels = QVBoxLayout()

        vbox_labels.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.Minimum, QSizePolicy.Minimum)) #for alignment

        PROG_LABELS = ["PCOM", "BCOM", "PM", "BA", "GLM", "FS", "DXD", "BK"]

        for label in PROG_LABELS:
            program = w.label(w.snow_reg, label)
            vbox_labels.addWidget(program)

        return vbox_labels


    def create_term_inputs(self, inputs, header_text):
        vbox = QVBoxLayout()

        for n in range(len(PROG_LABELS)):
            input_box = w.spin_box(w.glass, 0, 1000)
            input_box.setMinimumWidth(50)
            inputs.addWidget(input_box)

        inputs.setAlignment(Qt.AlignLeft)

        header = w.label(w.snow_reg, header_text)

        vbox.addWidget(header)
        vbox.addLayout(inputs)

        return vbox


    '''
    Helper Functions

    # The following section will be for action events
    # or functions that are called repeatedly
    # after the initial startup
    '''

    def update_class_combos(self):
        self.classroom_list.clear()
        self.select_room.clear()

        db = help_funcs.check_path("database\database.db")  # database.db file path
        connection = create_connection(db)
        db_classes = readClassroomItem(connection, "%")
        close_connection(connection)

        for each_class in range(len(db_classes)):

            if db_classes[each_class][2] == 1:
                self.select_room.addItem(db_classes[each_class][0] + f" Capacity: [{db_classes[each_class][1]}] (LAB)")
                self.classroom_list.addItem(db_classes[each_class][0] + f" Capacity: [{db_classes[each_class][1]}] (LAB)")
            else:
                self.select_room.addItem(db_classes[each_class][0] + f" Capacity: [{db_classes[each_class][1]}]")
                self.classroom_list.addItem(db_classes[each_class][0] + f" Capacity: [{db_classes[each_class][1]}]")


    '''
    Reset function
    '''
    def reset_table(self):
        # Use this to populate table with values to allow
        # Background colouring

        rows = self.main_table.rowCount()
        columns = self.main_table.columnCount()

        # necessary to display colour codes correctly
        self.main_table.setStyleSheet("background-color: None; color: #4f4f4f")

        for row in range(rows):
            for column in range(columns):
                placeholder = QTableWidgetItem()
                placeholder.setTextAlignment(Qt.AlignCenter)
                placeholder.setBackground(QtGui.QColor('#5e869c'))
                self.main_table.setItem(row, column, placeholder)
                self.main_table.removeCellWidget(row, column)


    def retrieve_term_inputs(self, layout):

        '''
        Used to read values from the term input fields.
        Will return a list of the inputs
        Always in the order of:
        PCOM
        BCOM
        PM
        BA
        GLM
        FS
        DXD
        BK
        '''
        input_fields = layout.count()
        list_val = []

        for each_field in range(input_fields):
            list_val.append(layout.itemAt(each_field).widget().value())

        return list_val

    # weekday is 0 or 2, depending on if its monday or wednesday respectively
    # eventually 1 and 3 for tuesday, thursday
    def show_schedule(self, lecture_list, weekday):

        global COLOUR_INDEX

        course = ""
        cohort_displayed = 0

        font = QFont()
        font.setPointSize(11)

        for cell in range(self.main_table.rowCount()):

            if lecture_list[cell] == "":
                continue

            # If this cell is the same name as the course from the previous cell,
            # Consider it part of the same block
            elif lecture_list[cell] != "" and course == lecture_list[cell]:
                self.main_table.item(cell,weekday).setBackground(QtGui.QColor(COURSE_COLOUR[course]))
                side_fill = QLabel()
                side_fill.setStyleSheet("border: solid white;"
                                        "border-width : 0px 2px 0px 2px;")

                if cell + 1 <= 26 and lecture_list[cell + 1] == "":
                    side_fill.setStyleSheet("border: solid white;"
                                            "border-width : 0px 2px 2px 2px;")

                if not cohort_displayed:
                    side_fill.setText(name[1])
                    side_fill.setAlignment(Qt.AlignCenter)
                    side_fill.setFont(font)
                    cohort_displayed = 1


                self.main_table.setCellWidget(cell, weekday, side_fill)

            # The course listed is a new one, and must be given a new colour + block
            else:
                course = lecture_list[cell]
                name = course.split("-")
                cohort_displayed = 0
                label_fill = QLabel(name[0])
                label_fill.setFont(font)
                label_fill.setAlignment(Qt.AlignCenter)
                label_fill.setStyleSheet("border: solid white;"
                                         "border-width : 2px 2px 0px 2px;")

                if cell + 1 <= 26 and lecture_list[cell + 1] != course:
                    label_fill.setStyleSheet("border: solid white;"
                                            "border-width : 0px 2px 2px 2px;")

                if course not in COURSE_COLOUR.keys():
                    COLOUR_INDEX += 1
                    if COLOUR_INDEX == len(BG_COLOURS):
                        COLOUR_INDEX = 0
                    COURSE_COLOUR[course] = BG_COLOURS[COLOUR_INDEX]
                self.main_table.setCellWidget(cell, weekday, label_fill)
                self.main_table.item(cell, weekday).setBackground(QtGui.QColor(COURSE_COLOUR[course]))


    def add_ghost_rooms(self):

        imports.fillClassrooms.fillClassrooms(SEM[self.pick_semester.currentText()])

        db = help_funcs.check_path("database\database.db")  # database.db file path
        connection = create_connection(db)
        if (len(readClassroomItem(connection, "ghost%")) != 0):
            ghost_rooms = QMessageBox(QMessageBox.Warning, "Insufficient Room Space", 
                                      "Be advised that additional rooms are required.")
            ghost_rooms.setStyleSheet("color: black")
            ghost_rooms.exec()

        close_connection(connection)


    '''
    Action Event functions
    '''
    def back_week(self):
        global CORE_SCHEDULE, PROG_SCHEDULE, WEEK


        # Can't go below week 1
        if WEEK == 1:
            return
        WEEK -= 1

        self.reset_table()
        # Get the lists for each day
        try:
            core = CORE_SCHEDULE[WEEK]
            prog = PROG_SCHEDULE[WEEK]

            monday = core[0]
            wednesday = core[1]
            tuesday = prog[0]
            thursday = prog[1]
            self.show_schedule(monday,0)
            self.show_schedule(tuesday, 1)
            self.show_schedule(wednesday, 2)
            self.show_schedule(thursday, 3)
            self.week_label.setText("Week of " + WEEK_DISPLAY_DATE[WEEK] + "\nWeek " + str(WEEK))
        except:
            return


    def forward_week(self):
        global CORE_SCHEDULE, PROG_SCHEDULE, WEEK


        # Only 13 weeks in a semester
        if WEEK == 13:
            return
        WEEK += 1

        self.reset_table()
        # Get the lists for each day
        try:
            core = CORE_SCHEDULE[WEEK]
            prog = PROG_SCHEDULE[WEEK]

            monday = core[0]
            wednesday = core[1]
            tuesday = prog[0]
            thursday = prog[1]
            self.show_schedule(monday,0)
            self.show_schedule(tuesday, 1)
            self.show_schedule(wednesday, 2)
            self.show_schedule(thursday, 3)
            self.week_label.setText("Week of " + WEEK_DISPLAY_DATE[WEEK] + "\nWeek " + str(WEEK))
        except:
            return

    # The following 2 functions are the same as forward week, but for
    # the cohort tab
    def forward_week_cohort(self):
        global CORE_SCHEDULE_COHORTS, PROG_SCHEDULE_COHORTS, WEEK_COHORTS


        # Only 13 weeks in a semester
        if WEEK_COHORTS == 13:
            return
        WEEK_COHORTS += 1

        self.reset_table_cohort()
        # Get the lists for each day
        try:
            core = CORE_SCHEDULE_COHORTS[WEEK_COHORTS]
            prog = PROG_SCHEDULE_COHORTS[WEEK_COHORTS]

            monday = core[0]
            wednesday = core[1]
            tuesday = prog[0]
            thursday = prog[1]
            self.show_schedule_cohorts(monday,0)
            self.show_schedule_cohorts(tuesday, 1)
            self.show_schedule_cohorts(wednesday, 2)
            self.show_schedule_cohorts(thursday, 3)
            self.cohort_week_label.setText("Week of " + WEEK_DISPLAY_DATE[WEEK_COHORTS] + "\nWeek " + str(WEEK_COHORTS))
        except:
            return
        

    def back_week_cohort(self):
    
        global CORE_SCHEDULE_COHORTS, PROG_SCHEDULE_COHORTS, WEEK_COHORTS

        # Cant go below week 1
        if WEEK_COHORTS == 1:
            return
        WEEK_COHORTS -= 1

        self.reset_table_cohort()
        # Get the lists for each day
        try:
            core = CORE_SCHEDULE_COHORTS[WEEK_COHORTS]
            prog = PROG_SCHEDULE_COHORTS[WEEK_COHORTS]

            monday = core[0]
            wednesday = core[1]
            tuesday = prog[0]
            thursday = prog[1]
            self.show_schedule_cohorts(monday,0)
            self.show_schedule_cohorts(tuesday, 1)
            self.show_schedule_cohorts(wednesday, 2)
            self.show_schedule_cohorts(thursday, 3)
            self.cohort_week_label.setText("Week of " + WEEK_DISPLAY_DATE[WEEK_COHORTS] + "\nWeek " + str(WEEK_COHORTS))
        except:
            return
        

    def create_schedule(self):
        global WEEK, CORE_SCHEDULE, PROG_SCHEDULE, CORE_DAY, PROG_DAY, ROOM, COURSE_COLOUR, COHORT_COURSE_COLOUR\
            ,CREATE_SCHEDULE_CLICKED, WEEK_DISPLAY_DATE, CORE_END_DATES, PROG_END_DATES, CORE_HOLIDAYS, PROG_HOLIDAYS
        # Reset values
        CORE_DAY = 0
        PROG_DAY = 0
        CORE_SCHEDULE.clear()
        PROG_SCHEDULE.clear()
        self.cohort_tab_combo.clear()
        WEEK = 1
        ROOM = self.select_room.currentText()

        if ROOM.find("(LAB)") != -1:
            ROOM = ROOM[:ROOM.find(" ")].strip() + " (LAB)"
            self.create_schedule_base(1)
        else:
            ROOM = ROOM[:ROOM.find(" ")].strip()
            self.create_schedule_base(0)


        # Clear out the lectures table
        # Pass in student numbers to db
        # Then calculate ghost rooms
        # Then parse the schedule.
        self.clear_lectures()
        self.pass_stu_num_db()
        self.add_ghost_rooms()

        # Loading message
        load_font = QFont()
        load_font.setPointSize(40)
        load_msg = QSplashScreen()
        load_msg.setFixedSize(400,200)
        load_msg.setStyleSheet(w.glass.specs)
        load_msg.setFont(load_font)
        load_msg.show()
        load_msg.showMessage("Loading...", 0, QColor(255,255,255))

        self.update_class_combos()

        #TODO Get export value
        export_confirm = self.export_check.checkState()

        schedule_info = imports.schedulers.core_scheduler.get_sched(SEM[self.pick_semester.currentText()], export_confirm)
        prog_schedule_info = imports.schedulers.program_scheduler.get_sched(SEM[self.pick_semester.currentText()], export_confirm)

        week_starts = schedule_info["week starts"]

        CORE_HOLIDAYS = schedule_info["holidays"]
        PROG_HOLIDAYS = prog_schedule_info["holidays"]

        CORE_END_DATES = schedule_info["last days"]
        PROG_END_DATES = prog_schedule_info["last days"]

        self.cohort_tab_combo.addItems(schedule_info["cohorts"])
        self.cohort_tab_combo.addItems(prog_schedule_info["cohorts"])

        index = 0
        for week_start in range(1, len(week_starts)):
            WEEK_DISPLAY_DATE[week_start] = week_starts[index].strftime("%Y-%m-%d")
            index+=1

        self.week_label.setText("Week of " + WEEK_DISPLAY_DATE[1] + "\nWeek " + str(WEEK))

        random.shuffle(BG_COLOURS)
        COURSE_COLOUR.clear()
        COHORT_COURSE_COLOUR.clear()
        self.reset_table()

        # All lecture items should now be recorded in the dictionaries
        self.get_lecture_items()

        # Get the lists for each day
        core = CORE_SCHEDULE[WEEK]
        prog = PROG_SCHEDULE[WEEK]

        monday = core[0]
        wednesday = core[1]
        tuesday = prog[0]
        thursday = prog[1]
        self.show_schedule(monday,0)
        self.show_schedule(tuesday, 1)
        self.show_schedule(wednesday, 2)
        self.show_schedule(thursday, 3)

        #close load message
        load_msg.close()

        #Success pop-up
        success = QMessageBox(QMessageBox.Information, "Success", 
                                      "Schedule has been created.")
        success.setStyleSheet("color: black")
        success.exec()

        CREATE_SCHEDULE_CLICKED = 1

    # Action event for creating template file
    def create_template(self):

        template = Workbook()

        sheet = template.active
        sheet.title = "Student Numbers"

        fields = ["Term", "Program", "Students"]

        sheet.append(fields)

        programs = ["PCOM", "BCOM", "PM", "BA", "GLM", "FS", "DXD", "BK"]

        for term in range (1, 4):
            for program in range(len(programs)):
                sheet.append([term, programs[program], 0])

        template.save("Student_Number_Inputs_Template.xlsx")
        template.close()


    # Action event for the choose file button
    def choose_file(self):

        chosen_file = QFileDialog.getOpenFileName(
            self,
            "Choose a file",
            os.getcwd(),
            "Excel Workbook files (*.xlsx)"  # Filters to specified file types
        )

        if chosen_file[0] == "":
            self.file_label.setText("No File Chosen")
            self.file_path = ""
        else:
            self.file_path = chosen_file[0]
            tokens = chosen_file[0].split("/")
            self.file_label.setText(tokens[-1])

    # Action event to load student input numbers
    def load_student_numbers(self):
        """
        Supports two excel templates.
        A template generated by the program, with total students per program.
        Client requested template, per student ID with program and term information.
        """
        try:
            stu_numbers = load_workbook(filename=self.file_path)

            try:
                sheet = stu_numbers.active

                #----------------------------------------
                # Template excel. Term, Program, Students
                if sheet["A1"].value == "Term":
                    term_1 = sheet["C2":"C9"]
                    term_2 = sheet["C10":"C17"]
                    term_3 = sheet["C18":"C25"]

                    for each_field in range(self.term_1_inputs.count()):
                        self.term_1_inputs.itemAt(each_field).widget().setValue(term_1[each_field][0].value)

                    for each_field in range(self.term_2_inputs.count()):
                        self.term_2_inputs.itemAt(each_field).widget().setValue(term_2[each_field][0].value)

                    for each_field in range(self.term_3_inputs.count()):
                        self.term_3_inputs.itemAt(each_field).widget().setValue(term_3[each_field][0].value)

                    stu_numbers.close()
                #----------------------------------------
                # Client requested excel. Student ID, Student Name, Program, Term
                elif sheet["A1"].value == "Student ID":
                    i = 2

                    program_counts = {"PCOM1": 0, "PCOM2": 0, "PCOM3": 0,
                                      "BCOM1": 0, "BCOM2": 0, "BCOM3": 0,
                                      "PM1":   0, "PM2":   0, "PM3":   0,
                                      "BA1":   0, "BA2":   0, "BA3":   0,
                                      "GLM1":  0, "GLM2":  0, "GLM3":  0,
                                      "FS1":   0, "FS2":   0, "FS3":   0,
                                      "DXD1":  0, "DXD2":  0, "DXD3":  0,
                                      "BK1":   0, "BK2":   0, "BK3":   0}

                    while sheet[f"A{i}"].value:

                        program  = sheet[f"C{i}"].value
                        term = str(sheet[f"D{i}"].value)
                        
                        program_counts[program + term] += 1

                        i += 1

                    keys = list(program_counts.keys())
                    for i in range(self.term_1_inputs.count()):
                        self.term_1_inputs.itemAt(i).widget().setValue(program_counts[keys[3*i + 0]])

                    for i in range(self.term_2_inputs.count()):
                        self.term_2_inputs.itemAt(i).widget().setValue(program_counts[keys[3*i + 1]])

                    for i in range(self.term_3_inputs.count()):
                        self.term_3_inputs.itemAt(i).widget().setValue(program_counts[keys[3*i + 2]])

                
                else: #pass
                    readErr = QMessageBox(QMessageBox.Warning, "Error!", 
                                        "Bad template. Please re-create")
                    readErr.setStyleSheet("color: black")
                    readErr.exec()
                    #print("Bad template!")

            except: #pass
                readErr = QMessageBox(QMessageBox.Warning, "Error Reading Values", 
                                        "Please Retry.")
                readErr.setStyleSheet("color: black")
                readErr.exec()
                #print("Error reading values")  # add error message here eventually

        except: #pass
            openErr = QMessageBox(QMessageBox.Warning, "Error Opening File", 
                                      "Please Retry.")
            openErr.setStyleSheet("color: black")
            openErr.exec()      
           # print("Error Opening File")# Maybe put an actual error message here eventually about opening files


    def pass_stu_num_db(self):

        term_1 = self.retrieve_term_inputs(self.term_1_inputs)
        term_2 = self.retrieve_term_inputs(self.term_2_inputs)
        term_3 = self.retrieve_term_inputs(self.term_3_inputs)

        programs = ["PCOM", "BCOM", "PM", "BA", "GLM", "FS", "DXD", "BK"]

        try:
            db = help_funcs.check_path("database\database.db")  # database.db file path
            connection = create_connection(db)

            # Clear table
            deleteStudentItem(connection, "%", "%")

            # Takes the currently input numbers, and adds them to the DB.
            for each_input in range(8):

                addStudentItem(connection, programs[each_input], 1, term_1[each_input])
                addStudentItem(connection, programs[each_input], 2, term_2[each_input])
                addStudentItem(connection, programs[each_input], 3, term_3[each_input])

            close_connection(connection)

        except:
            #print("Could not read database")
            close_connection(connection)

        return


    # Retrieve student items saved in db
    def load_db_stu_num(self):

        try:
            db = help_funcs.check_path("database\database.db")  # database.db file path
            connection = create_connection(db)

            term_1 = readStudentItem(connection, '%', 1)
            term_2 = readStudentItem(connection, '%', 2)
            term_3 = readStudentItem(connection, '%', 3)

            # Takes the currently input numbers, and adds them to the DB.
            for field in range(8):

                self.term_1_inputs.itemAt(field).widget().setValue(term_1[field][2])
                self.term_2_inputs.itemAt(field).widget().setValue(term_2[field][2])
                self.term_3_inputs.itemAt(field).widget().setValue(term_3[field][2])

            close_connection(connection)

        except:
            #print("Could not read database")
            close_connection(connection)

        return

    '''
    clears out the lecture table
    to prevent bloating
    and incorrect schedules
    '''
    def clear_lectures(self):
        try:
            db = help_funcs.check_path("database\database.db")  # database.db file path
            connection = create_connection(db)
            # Clear table
            deleteLectureItem_UI(connection)

        except:
            print("Could not read database")

        close_connection(connection)

    # Get the schedule for a specified semester, and fill the dictionary
    # With the 4-tuple lists for each week
    def get_lecture_items(self):

        global CORE_SCHEDULE, PROG_SCHEDULE, CORE_DAY, PROG_DAY, ROOM, CORE_END_DATES, PROG_END_DATES\
            , CORE_HOLIDAYS, PROG_HOLIDAYS
        
        core_week_list = []
        prog_week_list = []

        core_cross_check = []
        prog_cross_check = []

        core_last_known_sched = [""] *26
        prog_last_known_sched = [""] *26

        db = help_funcs.check_path("database\database.db")  # database.db file path
        connection = create_connection(db)

        # Get classes set on day 0
        core_lectures_in_week = readLectureItem_UI(connection, ROOM, 0, 1)
        prog_lectures_in_week = readLectureItem_UI(connection, ROOM, 0, 0)

        # Recall that day 1 for Core is monday, Day 1 for Prog is Tuesday
        # 13 weeks in a semester
        for weeks in range(1, 14):

            # Create a list for each day (2), (26 slots) for each list to correspond for each time
            # Recall that CORE_DAY and  PROG DAY are independant
            # i.e. An odd COre Day / PRog day = Monday / Tuesday, even = Wednesday/ thursday
            for each_day in range(1, 3):

                core_pass = []
                prog_pass = []

                core_lectures_in_week.extend(readLectureItem_UI(connection, ROOM, CORE_DAY, 1))
                prog_lectures_in_week.extend(readLectureItem_UI(connection, ROOM, PROG_DAY, 0))

                # Append to overall list of lectures. Compare against dictionary for end dates
                # Remove entry if it is past the end date.
                core_cross_check += core_lectures_in_week
                for course in core_cross_check:
                    if not CORE_DAY > CORE_END_DATES.get(course[0], 0):
                        core_pass.append(course)

                prog_cross_check += prog_lectures_in_week
                for course in prog_cross_check:
                    if not PROG_DAY > PROG_END_DATES.get(course[0], 0):
                        prog_pass.append(course)

                if CORE_DAY in CORE_HOLIDAYS:
                    core_last_known_sched = [""] * 26
                else:
                    core_last_known_sched = self.convert_to_list(core_pass)

                if PROG_DAY in PROG_HOLIDAYS:
                    prog_last_known_sched = [""] * 26
                else:
                    prog_last_known_sched = self.convert_to_list(prog_pass)

                core_week_list.append(core_last_known_sched)
                prog_week_list.append(prog_last_known_sched)

                CORE_DAY += 1
                PROG_DAY += 1
                
            # All weeks should be account for in the dictionaries now
            # Deepcopies must be made, since python does by reference
            CORE_SCHEDULE[weeks] = copy.deepcopy(core_week_list)
            PROG_SCHEDULE[weeks] = copy.deepcopy(prog_week_list)
            # Clear the list of lists
            core_week_list.clear()
            prog_week_list.clear()

        close_connection(connection)


    def get_cohort_lecture_items(self):

        global CORE_SCHEDULE_COHORTS, PROG_SCHEDULE_COHORTS, CORE_DAY, PROG_DAY, COHORT_CHOSEN, CORE_END_DATES\
            , PROG_END_DATES, CORE_HOLIDAYS, PROG_HOLIDAYS

        core_week_list = []
        prog_week_list = []

        core_cross_check = []
        prog_cross_check = []

        core_last_known_sched = [""] * 26
        prog_last_known_sched = [""] * 26

        db = help_funcs.check_path("database\database.db")  # database.db file path
        connection = create_connection(db)

        core_lectures_in_week = readLectureItem_UI_cohorts(connection, COHORT_CHOSEN, 0, 1)
        prog_lectures_in_week = readLectureItem_UI_cohorts(connection, COHORT_CHOSEN, 0, 0)

        # Recall that day 1 for Core is monday, Day 1 for Prog is Tuesday
        # 13 weeks in a semester
        for weeks in range(1, 14):

            # Create a list for each day (2), (26 slots) for each list to correspond for each time
            # Recall that CORE_DAY and  PROG DAY are independant
            # i.e. An odd COre Day / PRog day = Monday / Tuesday, even = Wednesday/ thursday
            for each_day in range(1, 3):

                core_pass = []
                prog_pass = []

                core_lectures_in_week.extend(readLectureItem_UI_cohorts(connection, COHORT_CHOSEN, CORE_DAY, 1))
                core_lectures_in_week.extend(readLectureItem_UI_cohorts_online(connection, COHORT_CHOSEN[:-2], CORE_DAY, 1))

                prog_lectures_in_week.extend(readLectureItem_UI_cohorts(connection, COHORT_CHOSEN, PROG_DAY, 0))
                prog_lectures_in_week.extend(readLectureItem_UI_cohorts_online(connection, COHORT_CHOSEN[:-2], PROG_DAY, 0))

                # Append to overall list of lectures. Compare against dictionary for end dates
                # Remove entry if it is past the end date.
                core_cross_check += core_lectures_in_week
                for course in core_cross_check:
                    if not CORE_DAY > CORE_END_DATES.get(course[0], 0):
                        core_pass.append(course)

                prog_cross_check += prog_lectures_in_week
                for course in prog_cross_check:
                    if not PROG_DAY > PROG_END_DATES.get(course[0], 0):
                        prog_pass.append(course)

                # Checking if there was any new differences in schedule
                # if not, then simply add the last known schedule since it hasnt changed.
                if CORE_DAY in CORE_HOLIDAYS:
                    core_last_known_sched = [""] * 26
                else:
                    core_last_known_sched = self.cohort_convert_to_list(core_pass)

                if PROG_DAY in PROG_HOLIDAYS:
                    prog_last_known_sched = [""] * 26
                else:
                    prog_last_known_sched = self.cohort_convert_to_list(prog_pass)

                core_week_list.append(core_last_known_sched)
                prog_week_list.append(prog_last_known_sched)

                CORE_DAY += 1
                PROG_DAY += 1

            # All weeks should be account for in the dictionaries now
            # Deepcopies must be made, since python does by reference
            CORE_SCHEDULE_COHORTS[weeks] = copy.deepcopy(core_week_list)
            PROG_SCHEDULE_COHORTS[weeks] = copy.deepcopy(prog_week_list)
            # Clear the list of lists
            core_week_list.clear()
            prog_week_list.clear()

        close_connection(connection)

    # Creates a list using the data pulled from the DB
    # The list matches what should be displayed in the UI.
    def convert_to_list(self, db_pull):
        global TIMES
        day_sched = [""] * 26

        for each_lecture in range(len(db_pull)):
            start_in_list = TIMES[db_pull[each_lecture][9]]
            slots_needed = int(db_pull[each_lecture][6] / .5)

            for each_slot in range(start_in_list, start_in_list + slots_needed):
                day_sched[each_slot] = db_pull[each_lecture][0] + "-" + db_pull[each_lecture][2]

        return day_sched


    def add_edit_classroom(self):
        db = help_funcs.check_path("database\database.db")  # database.db file path
        connection = create_connection(db)

        try:

            # Adds a classroom if it does not exist in database currently.
            # If it does, treat is as an edit (i.e. remove it from db, then add it fresh)

            classroom_id = self.class_id.text().strip()
            classroom_id = classroom_id.replace(" ","-")
            wanted_class = readClassroomItem(connection, classroom_id)
            lab = self.class_lab.checkedButton().text()

            if classroom_id.isspace() or classroom_id == "":
                close_connection(connection)
                return

            val = 0

            if (lab == "Lab"):
                val = 1
            if(len(wanted_class) == 1):
                deleteClassroomItem(connection, classroom_id)

            new_room = Classroom(classroom_id, self.class_capacity.value(), val)
            addClassroomItem(connection, new_room)

            # Must close connection to update db before updating comboboxes
            close_connection(connection)

            # Update the combobox
            self.update_class_combos()
            success = QMessageBox(QMessageBox.Information, "Success", 
                                      "Room has been added or modified.")
            success.setStyleSheet("color: black")
            success.exec()


        except:
            close_connection(connection)
            failure = QMessageBox(QMessageBox.Error, "Failure", 
                                      "Room has been not added or modified.")
            failure.setStyleSheet("color: black")
            failure.exec()


    def remove_classroom(self):
        db = help_funcs.check_path("database\database.db")  # database.db file path
        connection = create_connection(db)

        try:

            # Remove the classroom from the DB
            classroom = self.classroom_list.currentText()
            # Remove all text past the room name
            index = classroom.find(" ")
            classroom = classroom[:index].strip()

            deleteClassroomItem(connection, classroom)

            close_connection(connection)

            # Update the combobox
            self.update_class_combos()
            success = QMessageBox(QMessageBox.Information, "Success", 
                                      "Room has been deleted.")
            success.setStyleSheet("color: black")
            success.exec()


        except:
            close_connection(connection)
            failure = QMessageBox(QMessageBox.Error, "Failure", 
                                      "Room has been not deleted.")
            failure.setStyleSheet("color: black")
            failure.exec()


    def save_course(self):
        db = help_funcs.check_path("database\database.db")  # database.db file path
        connection = create_connection(db)

        try:

            if self.courses_edit_new.checkedButton().text() == "New Course":
                course_id = self.course_id.text().strip()
            else:
                course_id = self.courses.currentText().strip()
                deleteProgramItem_UI(connection, course_id)

            if course_id.isspace() or course_id == "":
                close_connection(connection)
                return

            wanted_course = readCourseItem(connection, course_id)

            if(len(wanted_course) == 1):
                deleteCourseItem(connection, course_id)

            term_hours = self.course_term_hours.value()
            term = self.course_term.value()
            duration = self.course_duration.value()
            core = self.course_core.isChecked()
            online = self.course_online.isChecked()
            lab = self.course_lab.isChecked()
            pre_reqs = self.course_pre_reqs

            new_course = Course(course_id, 1, term_hours, term, duration, core, online,lab, pre_reqs)
            addCourseItem(connection, new_course)
            addProgramItem_UI(connection, self.course_program.currentText(), course_id)

            close_connection(connection)
            # Update the combobox
            self.update_course_combos()
            success = QMessageBox(QMessageBox.Information, "Success", 
                                      "Task completed successfully.")
            success.setStyleSheet("color: black")
            success.exec()

        except:
            #pass
            addErr = QMessageBox(QMessageBox.Warning, "Error Adding Course", 
                                    "Please Retry.")
            addErr.setStyleSheet("color: black")
            addErr.exec()
            #print("error adding course")


    def update_course_combos(self):
        db = help_funcs.check_path("database\database.db")  # database.db file path
        connection = create_connection(db)
        db_courses = readCourseItem(connection, "%")
        close_connection(connection)

        course_name_list = []

        for each_course in range(len(db_courses)):
            course_name_list.append(db_courses[each_course][0])

        self.courses.clear()
        self.courses.addItems(course_name_list)
        self.course_pre_req_selector.clear()
        self.course_pre_req_selector.addItems(course_name_list)


    def show_hide_course(self):

        if self.courses_edit_new.checkedButton().text() == "New Course":
            self.course_id.show()
            self.courses.hide()
        else:
            self.course_id.hide()
            self.courses.show()


    def add_pre_req(self):
        self.course_pre_reqs.append(self.course_pre_req_selector.currentText())
        self.course_pre_reqs_label.setText(", ".join(self.course_pre_reqs))


    def clear_pre_reqs(self):
        self.course_pre_reqs.clear()
        self.course_pre_reqs_label.clear()


    '''
    Following functions change the schedule when selecting
    room
    '''
    def room_selector_show_schedule(self):

        if not CREATE_SCHEDULE_CLICKED:
            return

        global WEEK, ROOM, CORE_DAY, PROG_DAY, WEEK_DISPLAY_DATE
        WEEK = 1
        self.week_label.setText("Week of " + WEEK_DISPLAY_DATE[1] + "\nWeek " + str(WEEK))
        CORE_DAY = 1
        PROG_DAY = 1
        ROOM = self.select_room.currentText()
        if ROOM.find("(LAB)") != -1:
            ROOM = ROOM[:ROOM.find(" ")].strip() + " (LAB)"
            self.create_schedule_base(1)
        else:
            ROOM = ROOM[:ROOM.find(" ")].strip()
            self.create_schedule_base(0)

        random.shuffle(BG_COLOURS)
        COURSE_COLOUR.clear()
        self.reset_table()


        try:
            # All lecture items should now be recorded in the dictionaries
            self.get_lecture_items()

            # Get the lists for each day
            core = CORE_SCHEDULE[WEEK]
            prog = PROG_SCHEDULE[WEEK]

            monday = core[0]
            wednesday = core[1]
            tuesday = prog[0]
            thursday = prog[1]
            self.show_schedule(monday,0)
            self.show_schedule(tuesday, 1)
            self.show_schedule(wednesday, 2)
            self.show_schedule(thursday, 3)
        except:
            return

    '''
    # ------------------------------------------------------------------------
    
    the following sections will be the slightly editted functions to handle the cohort table
    
    #--------------------------------------------------------------------------
    
    '''
    def cohort_selector_show_schedule(self):

        global WEEK_COHORTS, COHORT_CHOSEN, CORE_DAY, PROG_DAY, COHORT_COURSE_TO_ROOM, COHORT_COURSE_COLOUR, WEEK_DISPLAY_DATE
        WEEK_COHORTS = 1
        self.cohort_week_label.setText("Week of " + WEEK_DISPLAY_DATE[1] + "\nWeek " + str(WEEK_COHORTS))
        CORE_DAY = 1
        PROG_DAY = 1
        COHORT_CHOSEN = self.cohort_tab_combo.currentText()
        COHORT_COURSE_TO_ROOM.clear()

        random.shuffle(BG_COLOURS)
        COHORT_COURSE_COLOUR.clear()
        self.reset_table_cohort()


        # All lecture items should now be recorded in the dictionaries
        self.get_cohort_lecture_items()

        # Get the lists for each day
        core = CORE_SCHEDULE_COHORTS[WEEK_COHORTS]
        prog = PROG_SCHEDULE_COHORTS[WEEK_COHORTS]

        monday = core[0]
        wednesday = core[1]
        tuesday = prog[0]
        thursday = prog[1]
        self.show_schedule_cohorts(monday, 0)
        self.show_schedule_cohorts(tuesday, 1)
        self.show_schedule_cohorts(wednesday, 2)
        self.show_schedule_cohorts(thursday, 3)

    # Creates a list using the COhort chosen
    # The list matches what should be displayed in the UI.
    def cohort_convert_to_list(self, db_pull):
        global TIMES, COHORT_COURSE_TO_ROOM
        day_sched = [""] * 26

        for each_lecture in range(len(db_pull)):
            start_in_list = TIMES[db_pull[each_lecture][9]]
            slots_needed = int(db_pull[each_lecture][6] / .5)

            for each_slot in range(start_in_list, start_in_list + slots_needed):
                day_sched[each_slot] = db_pull[each_lecture][0] + "~" + db_pull[each_lecture][3]

        return day_sched

    # Creating the basics of the cohort sched tab.
    # This is identical to the main_tab
    def make_cohort_sched_tab(self):

        cohort_table_box = QVBoxLayout(self)
        week_choose = QHBoxLayout(self)

        # Left/Right Navigation Arrows
        left =  w.push_button(w.coal,"←", self.back_week_cohort)
        right = w.push_button(w.coal,"→", self.forward_week_cohort)

        week_choose.addWidget(left)
        week_choose.addWidget(self.cohort_week_label)
        week_choose.addWidget(right)

        cohort_table_box.addWidget(self.cohort_tab_combo)
        cohort_table_box.addLayout(week_choose)
        cohort_table_box.addWidget(self.cohort_table)

        return cohort_table_box


    def create_cohorts_schedule_base(self):

        days = ["Monday", "Tuesday", "Wednesday", "Thursday"]
        times = []

        # Creation of all times from 8:00 AM to 5:00 PM to use as row headers

        first_time = datetime.datetime(year=2000, month=1, day=1, hour=8, minute=00)
        time_dif = datetime.timedelta(minutes=30)

        times.append(first_time.strftime("%I:%M %p"))
        new_time = first_time + time_dif

        # Has to be up to 8:30pm since we can't know if a room extends that far.
        for half_hour in range(25):
            times.append(new_time.strftime("%I:%M %p"))
            new_time = new_time + time_dif


        self.cohort_table.setColumnCount(4)
        self.cohort_table.setHorizontalHeaderLabels(days)

        self.cohort_table.setRowCount(len(times))
        self.cohort_table.setVerticalHeaderLabels(times)

        # Fill with empty items to change background colours later
        self.reset_table_cohort()


    def reset_table_cohort(self):
        # Use this to populate table with values to allow
        # Background colouring

        rows = self.cohort_table.rowCount()
        columns = self.cohort_table.columnCount()

        # necessary to display colour codes correctly
        self.cohort_table.setStyleSheet("background-color: None; color: #4f4f4f")

        for row in range(rows):
            for column in range(columns):
                placeholder = QTableWidgetItem()
                placeholder.setTextAlignment(Qt.AlignCenter)
                placeholder.setBackground(QtGui.QColor('#5e869c'))
                self.cohort_table.setItem(row, column, placeholder)
                self.cohort_table.removeCellWidget(row, column)


    def show_schedule_cohorts(self, lecture_list, weekday):

        global COLOUR_INDEX

        course = ""
        room_displayed = 0

        font = QFont()
        font.setPointSize(11)

        for cell in range(self.cohort_table.rowCount()):

            if lecture_list[cell] == "":
                continue

            # If this cell is the same name as the course from the previous cell,
            # Consider it part of the same block
            elif lecture_list[cell] != "" and course == lecture_list[cell]:
                self.cohort_table.item(cell,weekday).setBackground(QtGui.QColor(COHORT_COURSE_COLOUR[course]))
                side_fill = QLabel()
                side_fill.setStyleSheet("border: solid white;"
                                        "border-width : 0px 2px 0px 2px;")

                if cell + 1 <= 26 and lecture_list[cell + 1] == "":
                    side_fill.setStyleSheet("border: solid white;"
                                            "border-width : 0px 2px 2px 2px;")

                if not room_displayed:
                    side_fill.setText(name[1])
                    side_fill.setAlignment(Qt.AlignCenter)
                    side_fill.setFont(font)
                    room_displayed = 1


                self.cohort_table.setCellWidget(cell, weekday, side_fill)

            # The course listed is a new one, and must be given a new colour + block
            else:
                course = lecture_list[cell]
                room_displayed = 0
                name = course.split("~")
                label_fill = QLabel(name[0])
                label_fill.setFont(font)
                label_fill.setAlignment(Qt.AlignCenter)
                label_fill.setStyleSheet("border: solid white;"
                                         "border-width : 2px 2px 0px 2px;")

                if cell + 1 <= 26 and lecture_list[cell + 1] != course:
                    label_fill.setStyleSheet("border: solid white;"
                                            "border-width : 0px 2px 2px 2px;")

                if course not in COHORT_COURSE_COLOUR.keys():
                    COLOUR_INDEX += 1
                    if COLOUR_INDEX == len(BG_COLOURS):
                        COLOUR_INDEX = 0
                    COHORT_COURSE_COLOUR[course] = BG_COLOURS[COLOUR_INDEX]
                self.cohort_table.setCellWidget(cell, weekday, label_fill)
                self.cohort_table.item(cell, weekday).setBackground(QtGui.QColor(COHORT_COURSE_COLOUR[course]))